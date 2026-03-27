import sys, subprocess, os
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import threading, time, json
from datetime import datetime

# PIL opcional - sin PIL el logo usa canvas, con PIL muestra el logo real
try:
    from PIL import Image, ImageTk
    PIL_OK = True
except ImportError:
    PIL_OK = False

# ── CONFIGURACION GITHUB ──────────────────────────────────────────────────────
GITHUB_USER   = "wichoo2"
GITHUB_REPO   = "ingreso-masivo-xpress"
GITHUB_BRANCH = "main"
GITHUB_TOKEN  = ""   # dejar vacio si el repo es publico

# URLs base
_GH_RAW  = "https://raw.githubusercontent.com/{}/{}/{}/{{}}".format(
    GITHUB_USER, GITHUB_REPO, GITHUB_BRANCH)
_GH_VER  = _GH_RAW.format("version.json")

# ── SISTEMA DE ACTUALIZACIONES ────────────────────────────────────────────────
def _leer_version_local():
    try:
        base = os.path.dirname(os.path.abspath(__file__))
        ruta = os.path.join(base, "version.json")
        with open(ruta, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {"version": "0.0"}

def _consultar_version_github():
    """
    Consulta version.json en GitHub usando urllib (sin dependencias externas).
    Agrega timestamp anti-cache para siempre obtener la version mas reciente.
    Retorna el dict de version o None si hay error.
    """
    try:
        import urllib.request
        url = "{}?_={}".format(_GH_VER, int(time.time()))
        req = urllib.request.Request(url)
        req.add_header("Cache-Control", "no-cache, no-store, must-revalidate")
        req.add_header("Pragma", "no-cache")
        req.add_header("Expires", "0")
        if GITHUB_TOKEN:
            req.add_header("Authorization", "token {}".format(GITHUB_TOKEN))
        with urllib.request.urlopen(req, timeout=8) as r:
            return json.loads(r.read().decode("utf-8"))
    except Exception:
        return None

def _descargar_archivo_github(nombre_archivo):
    """Descarga un archivo desde GitHub y retorna su contenido como bytes."""
    try:
        import urllib.request
        url = "{}?_={}".format(_GH_RAW.format(nombre_archivo), int(time.time()))
        req = urllib.request.Request(url)
        if GITHUB_TOKEN:
            req.add_header("Authorization", "token {}".format(GITHUB_TOKEN))
        with urllib.request.urlopen(req, timeout=15) as r:
            return r.read()
    except Exception as e:
        return None

def _hay_actualizacion():
    """
    Compara la version local con GitHub.
    Retorna (hay_update: bool, version_remota: dict | None)
    """
    local   = _leer_version_local()
    remota  = _consultar_version_github()
    if remota is None:
        return False, None
    v_local  = str(local.get("version",  "0.0")).strip()
    v_remota = str(remota.get("version", "0.0")).strip()
    return v_local != v_remota, remota

def _aplicar_actualizacion(version_remota, callback_progreso=None, callback_log=None):
    """
    Descarga y reemplaza los archivos listados en version.json de GitHub.
    - Nunca toca config_local.py
    - Hace backup .bak de cada archivo antes de reemplazar
    - Actualiza version.json local al final
    Retorna (ok: bool, mensaje: str)
    """
    base     = os.path.dirname(os.path.abspath(__file__))
    archivos = version_remota.get("archivos", [])
    total    = len(archivos) + 1  # +1 por version.json al final

    def _log(msg):
        if callback_log:
            callback_log(msg)

    def _prog(n):
        if callback_progreso:
            callback_progreso(n / total)

    errores = []
    for i, nombre in enumerate(archivos):
        # Nunca actualizar config_local.py para no borrar rutas del usuario
        if nombre == "config_local.py":
            _log("  [SKIP] config_local.py — no se actualiza")
            _prog(i + 1)
            continue

        _log("  Descargando {}...".format(nombre))
        contenido = _descargar_archivo_github(nombre)
        if contenido is None:
            errores.append(nombre)
            _log("  [ERROR] No se pudo descargar {}".format(nombre))
            _prog(i + 1)
            continue

        ruta_dst = os.path.join(base, nombre)
        # Backup del archivo anterior
        if os.path.isfile(ruta_dst):
            try:
                with open(ruta_dst + ".bak", "wb") as f_bak:
                    with open(ruta_dst, "rb") as f_src:
                        f_bak.write(f_src.read())
            except Exception:
                pass

        # Escribir nuevo contenido
        try:
            with open(ruta_dst, "wb") as f:
                f.write(contenido)
            _log("  ✓ {}".format(nombre))
        except Exception as e:
            errores.append(nombre)
            _log("  [ERROR] {}: {}".format(nombre, e))

        _prog(i + 1)

    # Actualizar version.json local
    try:
        ruta_ver = os.path.join(base, "version.json")
        with open(ruta_ver, "w", encoding="utf-8") as f:
            json.dump(version_remota, f, ensure_ascii=False, indent=2)
        _log("  ✓ version.json actualizado a {}".format(
            version_remota.get("version")))
    except Exception as e:
        _log("  [ERROR] version.json: {}".format(e))

    _prog(total)

    if errores:
        return False, "Errores en: {}".format(", ".join(errores))
    return True, "OK"

def _mostrar_error(err):
    try:
        r = tk.Tk(); r.withdraw()
        messagebox.showerror("Error al iniciar", str(err))
        r.destroy()
    except: pass

def _check_deps():
    missing = []
    try: import openpyxl
    except ImportError: missing.append("openpyxl")
    if not missing:
        return

    # Buscar Python real — el _internal del launcher o el del sistema
    import shutil as _sh
    _base = os.path.dirname(os.path.abspath(__file__))
    _python = (
        os.path.join(_base, "_internal", "pythonw.exe") if os.path.isfile(
            os.path.join(_base, "_internal", "pythonw.exe")) else
        os.path.join(_base, "_internal", "python.exe") if os.path.isfile(
            os.path.join(_base, "_internal", "python.exe")) else
        _sh.which("python") or _sh.which("pythonw") or sys.executable
    )

    r = tk.Tk(); r.withdraw()
    ok = messagebox.askyesno("Dependencias",
        "Faltan modulos:\n\n  {}\n\nInstalar ahora?".format(
            ", ".join(missing)))
    r.destroy()
    if ok:
        for pkg in missing:
            subprocess.check_call([_python, "-m", "pip", "install", pkg])
    else:
        sys.exit(0)

_check_deps()

BASE_DIR    = os.path.dirname(os.path.abspath(__file__))
MAIN_PY     = os.path.join(BASE_DIR, "main_local.py")
TEST_PY     = os.path.join(BASE_DIR, "test.py")
INDEXAR_PY  = os.path.join(BASE_DIR, "indexar.py")
CONFIG_PY   = os.path.join(BASE_DIR, "config_local.py")
STATS_FILE  = os.path.join(BASE_DIR, "ultimo_resultado.json")
CONFIG_PASS = "7070"

# ── PALETA XPRESS ─────────────────────────────────────────────────────────────
BG      = "#0a0a0f"
BG2     = "#111118"
BG3     = "#18181f"
BG4     = "#1e1e28"
BG5     = "#242430"
BORDER  = "#2a2a38"
BORDER2 = "#353548"
ORANGE  = "#FF4500"
ORANGE2 = "#E03D00"
ORANGE3 = "#FF6B35"
ORANGEDK= "#8B2500"
WHITE   = "#FFFFFF"
WHITE2  = "#E8E8F0"
WHITE3  = "#A0A0B8"
WHITE4  = "#606078"
GREEN   = "#00E5A0"
RED     = "#FF4060"
YELLOW  = "#FFB020"
BLUE    = "#4090FF"
UI      = ("Segoe UI", 10)
UI_SM   = ("Segoe UI", 9)
UI_B    = ("Segoe UI", 10, "bold")
UI_LG   = ("Segoe UI", 14, "bold")
MONO    = ("Consolas", 9)

LOGO_B64 = (
    "iVBORw0KGgoAAAANSUhEUgAAAQAAAAEACAYAAABccqhmAAAACXBIWXMAAAsTAAALEwEAmpwY"
    "AAAF8WlUWHRYTUw6Y29tLmFkb2JlLnhtcAAAAAAAPD94cGFja2V0IGJlZ2luPSLvu78iIGlk"
    "PSJXNU0wTXBDZWhpSHpyZVN6TlRjemtjOWQiPz4gPHg6eG1wbWV0YSB4bWxuczp4PSJhZG9i"
    "ZTpuczptZXRhLyIgeDp4bXB0az0iQWRvYmUgWE1QIENvcmUgOS4xLWMwMDIgNzkuZjM1NGVm"
)

def cargar_stats():
    try:
        with open(STATS_FILE) as f: return json.load(f)
    except: return {"listo":0,"falta":0,"dup":0,"tiempo":None,"fecha":None}

def guardar_stats(l,f,d,t):
    with open(STATS_FILE,"w") as fp:
        json.dump({"listo":l,"falta":f,"dup":d,"tiempo":round(t,1),
                   "fecha":datetime.now().strftime("%d/%m/%Y %I:%M %p")},fp)

def fmt_hora():
    return datetime.now().strftime("%A %d de %B, %Y  |  %I:%M %p")

def cargar_logo_tk(size):
    if not PIL_OK:
        return None
    try:
        ruta = os.path.join(BASE_DIR, "xpress_logo.png")
        if os.path.isfile(ruta):
            img = Image.open(ruta).convert("RGBA").resize((size,size), Image.LANCZOS)
            return ImageTk.PhotoImage(img)
    except: pass
    return None

# ═══════════════════════════════════════════════════════════════════════════════
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Ingreso Masivo | Xpress El Salvador")
        self.geometry("1200x720")
        self.minsize(1000,600)
        self.configure(bg=BG)
        try:
            ico = cargar_logo_tk(32)
            if ico: self.iconphoto(True, ico)
        except: pass
        self._proceso_activo = False
        self._t_inicio       = None
        self._pulse_after    = None
        self._total_filas    = 0
        self._proc_filas     = 0
        self._eta_after      = None
        self._tiempos_fila   = []
        self._version_remota = None   # se llena si hay update disponible
        self._build_ui()
        self.bind("<F5>", lambda e: self._ejecutar())
        # Verificar actualizaciones en segundo plano al abrir
        threading.Thread(target=self._check_update_bg, daemon=True).start()

    # ── BUILD UI ──────────────────────────────────────────────────────────────
    def _build_ui(self):
        # Sidebar
        sb = tk.Frame(self, bg=BG2, width=250)
        sb.pack(side="left", fill="y")
        sb.pack_propagate(False)

        # Logo area con gradiente simulado
        logo_area = tk.Frame(sb, bg=BG2, pady=20)
        logo_area.pack(fill="x")

        # Canvas para el logo
        logo_cv = tk.Canvas(logo_area, width=64, height=64,
                            bg=BG2, highlightthickness=0)
        logo_cv.pack(pady=(10,0))

        # Intentar cargar logo real
        self._logo_img = cargar_logo_tk(56)
        if self._logo_img:
            logo_cv.create_image(32, 32, image=self._logo_img)
        else:
            # Fallback: dibujar aguila estilizada con canvas
            logo_cv.create_oval(4,4,60,60, fill=ORANGE, outline="")
            logo_cv.create_text(32,32, text="XP", fill=WHITE,
                               font=("Segoe UI",16,"bold"))

        # Linea decorativa naranja bajo el logo
        tk.Frame(logo_area, bg=ORANGE, height=2, width=40).pack(pady=(8,0))

        tk.Label(logo_area, text="INGRESO MASIVO", bg=BG2, fg=WHITE,
                 font=("Segoe UI",11,"bold")).pack(pady=(8,0))
        tk.Label(logo_area, text="Xpress El Salvador", bg=BG2, fg=WHITE3,
                 font=("Segoe UI",8)).pack()

        self._lbl_sb_estado = tk.Label(logo_area, text="● Listo",
                                        bg=BG2, fg=GREEN,
                                        font=("Segoe UI",8))
        self._lbl_sb_estado.pack(pady=(4,0))

        # Separador con acento naranja
        sep_frame = tk.Frame(sb, bg=BG2, pady=2)
        sep_frame.pack(fill="x", padx=16)
        tk.Frame(sep_frame, bg=BORDER2, height=1).pack(fill="x")
        tk.Frame(sep_frame, bg=ORANGE, height=1, width=30).pack(anchor="w")

        # Botones sidebar
        self._mk_sec(sb, "ACCIONES")
        self._sbtn(sb, "Ejecutar",       "F5  |  Procesar INGRESO_MASIVO",   ORANGE,  self._ejecutar, True)
        self._sbtn(sb, "Verificar",      "Comprobar archivos .xlsx",          BLUE,    self._verificar)
        self._sbtn(sb, "Indexar tiendas","Actualizar indice col E",           ORANGE3, self._indexar)
        self._btn_deshacer = self._sbtn(sb, "⟲ Deshacer",   "Revertir el ultimo proceso",        "#8B4513",   self._deshacer_proceso)

        tk.Frame(sb, bg=BORDER, height=1).pack(fill="x", padx=16, pady=6)
        self._mk_sec(sb, "REPORTES")
        self._sbtn(sb, "Ver FALTA",      "Filas sin tienda destino",          YELLOW,  self._ver_falta)
        self._sbtn(sb, "Diagnosticar",   "Por que no se ingresa una tienda",  RED,     self._abrir_diagnostico)
        self._sbtn(sb, "Cambiar rutas",  "Configurar archivos y carpetas",    WHITE4,  self._abrir_config)
        self._sbtn(sb, "Lista omisiones","Forzar ingreso por nombre manual",  ORANGE3, self._abrir_omisiones)
        self._sbtn(sb, "Lista negra",    "Tiendas bloqueadas permanentemente", RED,     self._abrir_blacklist)

        tk.Frame(sb, bg=BORDER, height=1).pack(fill="x", padx=16, pady=6)

        # Banner de actualizacion (oculto por defecto)
        self._update_banner = tk.Frame(sb, bg="#1a3a1a", cursor="hand2")
        self._update_lbl    = tk.Label(self._update_banner, text="",
                                        bg="#1a3a1a", fg=GREEN,
                                        font=("Segoe UI", 8, "bold"),
                                        wraplength=200, justify="left")
        self._update_lbl.pack(padx=10, pady=6)
        self._update_banner.bind("<Button-1>", lambda e: self._mostrar_dialogo_update())
        self._update_lbl.bind("<Button-1>",    lambda e: self._mostrar_dialogo_update())
        # No se hace pack todavía — se muestra solo si hay update

        # Info abajo del sidebar
        info = tk.Frame(sb, bg=BG2)
        info.pack(side="bottom", fill="x", padx=16, pady=14)
        ver_local = _leer_version_local().get("version", "?")
        self._lbl_ver = tk.Label(info,
                                  text="F5 ejecutar  |  v{}".format(ver_local),
                                  bg=BG2, fg=WHITE4, font=("Segoe UI",8))
        self._lbl_ver.pack(anchor="w")
        tk.Label(info, text="Xpress El Salvador 2026", bg=BG2,
                 fg=WHITE4, font=("Segoe UI",7)).pack(anchor="w", pady=(2,0))

        # ── Main panel ────────────────────────────────────────────────────────
        main = tk.Frame(self, bg=BG)
        main.pack(side="left", fill="both", expand=True)

        # Header con acento
        hdr = tk.Frame(main, bg=BG, padx=28, pady=16)
        hdr.pack(fill="x")
        left_hdr = tk.Frame(hdr, bg=BG)
        left_hdr.pack(side="left", fill="x", expand=True)

        # Linea acento naranja vertical
        tk.Frame(left_hdr, bg=ORANGE, width=3, height=36).pack(
            side="left", padx=(0,12))
        title_f = tk.Frame(left_hdr, bg=BG)
        title_f.pack(side="left")
        tk.Label(title_f, text="Panel de Control", bg=BG, fg=WHITE,
                 font=("Segoe UI",16,"bold")).pack(anchor="w")
        self._lbl_fecha = tk.Label(title_f, text=fmt_hora(),
                                   bg=BG, fg=WHITE3, font=("Segoe UI",9))
        self._lbl_fecha.pack(anchor="w")
        self._tick_fecha()

        # Stat cards — 5 tarjetas
        cf = tk.Frame(main, bg=BG, padx=28)
        cf.pack(fill="x", pady=(0,14))
        self._c_listo  = self._card(cf, "0",  "PROCESADOS", GREEN)
        self._c_falta  = self._card(cf, "0",  "FALTA",      RED)
        self._c_dup    = self._card(cf, "0",  "DUPLICADOS", YELLOW)
        self._c_tiempo = self._card(cf, "--", "SEGUNDOS",   ORANGE)
        self._c_fecha  = self._card(cf, "--", "ULTIMO",     WHITE3)
        for c in cf.winfo_children():
            c.pack(side="left", fill="x", expand=True, padx=(0,8))

        # ── Barra de progreso REAL (filas procesadas) ─────────────────────────
        pf_outer = tk.Frame(main, bg=BG, padx=28)
        pf_outer.pack(fill="x", pady=(0,6))

        # Barra superior: progreso real por filas
        pf_top = tk.Frame(pf_outer, bg=BG3,
                          highlightbackground=BORDER2, highlightthickness=1)
        pf_top.pack(fill="x", pady=(0,4))
        top_hdr = tk.Frame(pf_top, bg=BG3, padx=16, pady=8)
        top_hdr.pack(fill="x")
        tk.Label(top_hdr, text="Progreso", bg=BG3,
                 fg=WHITE3, font=("Segoe UI",8,"bold")).pack(side="left")
        self._lbl_pct_real = tk.Label(top_hdr, text="0%", bg=BG3,
                                       fg=ORANGE, font=("Segoe UI",8,"bold"))
        self._lbl_pct_real.pack(side="right")
        self._lbl_eta = tk.Label(top_hdr, text="ETA: --", bg=BG3,
                                  fg=BLUE, font=("Segoe UI",8,"bold"))
        self._lbl_eta.pack(side="right", padx=12)
        self._lbl_filas = tk.Label(top_hdr, text="0 / 0 filas",
                                    bg=BG3, fg=WHITE4, font=("Segoe UI",8))
        self._lbl_filas.pack(side="right", padx=10)

        sty = ttk.Style()
        sty.theme_use("clam")
        sty.configure("Real.Horizontal.TProgressbar",
                       troughcolor=BG2, background=ORANGE,
                       bordercolor=BORDER, lightcolor=ORANGE,
                       darkcolor=ORANGE2, thickness=8)
        self._prog_real = ttk.Progressbar(pf_top, style="Real.Horizontal.TProgressbar",
                                           mode="determinate", maximum=100)
        self._prog_real.pack(fill="x", padx=16, pady=(0,10))

        # Barra inferior: paso actual detallado
        pf_bot = tk.Frame(pf_outer, bg=BG3,
                          highlightbackground=BORDER2, highlightthickness=1)
        pf_bot.pack(fill="x")
        bot_hdr = tk.Frame(pf_bot, bg=BG3, padx=16, pady=8)
        bot_hdr.pack(fill="x")
        # Icono de fase
        self._lbl_fase_ico = tk.Label(bot_hdr, text="⏸", bg=BG3,
                                       fg=WHITE4, font=("Segoe UI",9))
        self._lbl_fase_ico.pack(side="left", padx=(0,6))
        self._lbl_actividad = tk.Label(bot_hdr, text="En espera — presiona F5 para iniciar",
                                        bg=BG3, fg=WHITE3, font=("Segoe UI",8))
        self._lbl_actividad.pack(side="left", fill="x", expand=True)
        self._lbl_elapsed = tk.Label(bot_hdr, text="", bg=BG3,
                                      fg=WHITE4, font=("Consolas",8))
        self._lbl_elapsed.pack(side="right")

        sty.configure("Act.Horizontal.TProgressbar",
                       troughcolor=BG2, background=ORANGE3,
                       bordercolor=BORDER, lightcolor=ORANGE3,
                       darkcolor=ORANGEDK, thickness=4)
        self._prog_act = ttk.Progressbar(pf_bot, style="Act.Horizontal.TProgressbar",
                                          mode="indeterminate")
        self._prog_act.pack(fill="x", padx=16, pady=(0,10))

        # ── Log area con panel redimensionable ───────────────────────────────
        log_outer = tk.Frame(main, bg=BG)
        log_outer.pack(fill="both", expand=True, padx=28, pady=(8,0))

        # PanedWindow para que el log sea redimensionable arrastrando
        paned = tk.PanedWindow(log_outer, orient="horizontal",
                               bg=BG, sashwidth=6,
                               sashrelief="flat", sashpad=2)
        paned.pack(fill="both", expand=True)

        # Panel izquierdo: log principal
        log_wrap = tk.Frame(paned, bg=BG3,
                            highlightbackground=BORDER2, highlightthickness=1)
        paned.add(log_wrap, minsize=400)

        log_hdr = tk.Frame(log_wrap, bg=BG3, padx=14, pady=8)
        log_hdr.pack(fill="x")
        # Indicador de actividad (punto pulsante)
        self._dot_cv = tk.Canvas(log_hdr, width=8, height=8,
                                  bg=BG3, highlightthickness=0)
        self._dot_cv.pack(side="left", padx=(0,6))
        self._dot = self._dot_cv.create_oval(0,0,8,8, fill=WHITE4, outline="")
        tk.Label(log_hdr, text="Actividad en tiempo real", bg=BG3,
                 fg=WHITE2, font=UI_B).pack(side="left")
        tk.Button(log_hdr, text="Limpiar", bg=BG3, fg=WHITE4,
                  font=("Segoe UI",8), relief="flat", cursor="hand2",
                  bd=0, activebackground=BG3, activeforeground=WHITE2,
                  command=self._limpiar_log).pack(side="right")
        tk.Frame(log_wrap, bg=BORDER, height=1).pack(fill="x")

        # Canvas scrollable para el log
        self._lcanvas = tk.Canvas(log_wrap, bg=BG3, highlightthickness=0)
        self._lsb     = tk.Scrollbar(log_wrap, orient="vertical",
                                      command=self._lcanvas.yview)
        self._lcanvas.configure(yscrollcommand=self._lsb.set)
        self._lsb.pack(side="right", fill="y")
        self._lcanvas.pack(side="left", fill="both", expand=True)
        self._lframe = tk.Frame(self._lcanvas, bg=BG3)
        self._lwin   = self._lcanvas.create_window(
            (0,0), window=self._lframe, anchor="nw")
        self._lframe.bind("<Configure>",
            lambda e: self._lcanvas.configure(
                scrollregion=self._lcanvas.bbox("all")))
        self._lcanvas.bind("<Configure>",
            lambda e: self._lcanvas.itemconfig(self._lwin, width=e.width))
        # Scroll con rueda del raton
        self._lcanvas.bind("<MouseWheel>",
            lambda e: self._lcanvas.yview_scroll(
                int(-1*(e.delta/120)), "units"))

        # Panel derecho: resumen + progreso por tienda
        right = tk.Frame(paned, bg=BG, width=220)
        paned.add(right, minsize=200)

        tk.Label(right, text="Ultimo proceso", bg=BG,
                 fg=WHITE2, font=UI_B).pack(anchor="w", padx=4, pady=(0,8))
        self._rframe = tk.Frame(right, bg=BG)
        self._rframe.pack(fill="x", padx=4)
        self._build_resumen()

        # Panel de progreso por tienda (reemplaza "Estado actual")
        tk.Frame(right, bg=BORDER, height=1).pack(fill="x", padx=4, pady=10)

        tiendas_hdr = tk.Frame(right, bg=BG)
        tiendas_hdr.pack(fill="x", padx=4, pady=(0,6))
        tk.Label(tiendas_hdr, text="Tiendas en proceso", bg=BG,
                 fg=WHITE2, font=UI_B).pack(side="left")
        self._lbl_tiendas_cnt = tk.Label(tiendas_hdr, text="",
                                          bg=BG, fg=WHITE4,
                                          font=("Segoe UI",7))
        self._lbl_tiendas_cnt.pack(side="right")

        # Canvas scrollable para las tarjetas de tienda
        tiendas_outer = tk.Frame(right, bg=BG)
        tiendas_outer.pack(fill="both", expand=True, padx=4)
        self._tiendas_canvas = tk.Canvas(tiendas_outer, bg=BG,
                                          highlightthickness=0)
        tiendas_vsb = tk.Scrollbar(tiendas_outer, orient="vertical",
                                    command=self._tiendas_canvas.yview)
        self._tiendas_canvas.configure(yscrollcommand=tiendas_vsb.set)
        tiendas_vsb.pack(side="right", fill="y")
        self._tiendas_canvas.pack(side="left", fill="both", expand=True)
        self._tiendas_frame = tk.Frame(self._tiendas_canvas, bg=BG)
        self._tiendas_win   = self._tiendas_canvas.create_window(
            (0, 0), window=self._tiendas_frame, anchor="nw")
        self._tiendas_frame.bind("<Configure>",
            lambda e: self._tiendas_canvas.configure(
                scrollregion=self._tiendas_canvas.bbox("all")))
        self._tiendas_canvas.bind("<Configure>",
            lambda e: self._tiendas_canvas.itemconfig(
                self._tiendas_win, width=e.width))
        self._tiendas_canvas.bind("<MouseWheel>",
            lambda e: self._tiendas_canvas.yview_scroll(
                int(-1*(e.delta/120)), "units"))

        # Dict de tarjetas activas: nombre_tienda -> {frame, canvas, fill, lbl_estado, lbl_cnt}
        self._tarjetas_tienda = {}

        # Footer
        footer = tk.Frame(main, bg=BG2, height=28)
        footer.pack(fill="x", side="bottom")
        footer.pack_propagate(False)
        _ver_actual = _leer_version_local().get("version", "?")
        tk.Label(footer, text="Ingreso Masivo  |  Xpress El Salvador  |  v{}".format(_ver_actual),
                 bg=BG2, fg=WHITE4,
                 font=("Segoe UI",8)).pack(side="left", padx=16, pady=6)
        self._lbl_footer_hora = tk.Label(footer, text="",
                                          bg=BG2, fg=WHITE4,
                                          font=("Segoe UI",8))
        self._lbl_footer_hora.pack(side="right", padx=16)
        self._tick_footer()

    # ── COMPONENTES ──────────────────────────────────────────────────────────
    def _mk_sec(self, p, txt):
        f = tk.Frame(p, bg=BG2, padx=16)
        f.pack(fill="x", pady=(6,2))
        tk.Label(f, text=txt, bg=BG2, fg=WHITE4,
                 font=("Segoe UI",7,"bold")).pack(anchor="w")

    def _sbtn(self, p, titulo, sub, color, cmd, destacado=False):
        bg_base = BG3 if destacado else BG2
        f = tk.Frame(p, bg=bg_base, cursor="hand2")
        f.pack(fill="x", padx=8, pady=2)

        # Borde izquierdo de color si es destacado
        if destacado:
            tk.Frame(f, bg=ORANGE, width=3).pack(side="left", fill="y")

        inn = tk.Frame(f, bg=bg_base, padx=12, pady=8)
        inn.pack(side="left", fill="x", expand=True)

        dot = tk.Canvas(inn, width=8, height=8,
                        bg=bg_base, highlightthickness=0)
        dot.pack(side="left", padx=(0,10))
        dot.create_oval(0,0,8,8, fill=color, outline="")

        tf = tk.Frame(inn, bg=bg_base)
        tf.pack(side="left", fill="x", expand=True)
        lc = ORANGE if destacado else WHITE2
        lt = tk.Label(tf, text=titulo, bg=bg_base, fg=lc,
                      font=UI_B if destacado else UI, anchor="w")
        lt.pack(anchor="w")
        ls = tk.Label(tf, text=sub, bg=bg_base, fg=WHITE4,
                      font=("Segoe UI",8), anchor="w")
        ls.pack(anchor="w")

        ws = [f, inn, dot, tf, lt, ls]
        def _e(e):
            for w in ws: w.configure(bg=BG4)
            dot.configure(bg=BG4)
        def _l(e):
            for w in ws: w.configure(bg=bg_base)
            dot.configure(bg=bg_base)
        def _c(e): cmd()
        for w in ws:
            w.bind("<Enter>", _e)
            w.bind("<Leave>", _l)
            w.bind("<Button-1>", _c)

    def _card(self, p, val, lbl, color):
        c = tk.Frame(p, bg=BG3,
                     highlightbackground=BORDER2, highlightthickness=1)
        # Borde superior de color
        tk.Frame(c, bg=color, height=2).pack(fill="x")
        lv = tk.Label(c, text=val, bg=BG3, fg=color,
                      font=("Segoe UI",22,"bold"))
        lv.pack(padx=14, pady=(10,2))
        tk.Label(c, text=lbl, bg=BG3, fg=WHITE4,
                 font=("Segoe UI",7,"bold")).pack(padx=14, pady=(0,12))
        c._lv = lv
        return c

    def _set_card(self, c, v): c._lv.configure(text=str(v))

    def _build_resumen(self):
        for w in self._rframe.winfo_children(): w.destroy()
        s = cargar_stats()
        rows = [
            ("Procesados", s["listo"],  GREEN),
            ("Falta",      s["falta"],  RED),
            ("Duplicados", s["dup"],    YELLOW),
            ("Segundos",   f"{s['tiempo']}s" if s["tiempo"] else "--", ORANGE),
            ("Ultima vez", s["fecha"] or "--", WHITE3),
        ]
        for lbl, val, col in rows:
            r = tk.Frame(self._rframe, bg=BG4,
                         highlightbackground=BORDER, highlightthickness=1)
            r.pack(fill="x", pady=2)
            tk.Label(r, text=lbl, bg=BG4, fg=WHITE4,
                     font=UI_SM, width=10, anchor="w").pack(
                         side="left", padx=10, pady=6)
            tk.Label(r, text=str(val), bg=BG4, fg=col,
                     font=UI_B).pack(side="right", padx=10)

    def _build_estado_actual(self):
        # Compatibilidad — ya no se usa pero puede llamarse desde código antiguo
        pass

    def _tienda_card_crear(self, nombre):
        """Crea una tarjeta de tienda en el panel derecho."""
        if nombre in self._tarjetas_tienda:
            return

        nombre_corto = nombre[:28] + "…" if len(nombre) > 28 else nombre

        card = tk.Frame(self._tiendas_frame, bg=BG3,
                        highlightbackground=BORDER2, highlightthickness=1)
        card.pack(fill="x", pady=2, padx=2)

        # Nombre
        tk.Label(card, text=nombre_corto, bg=BG3, fg=WHITE2,
                 font=("Segoe UI", 7, "bold"),
                 anchor="w").pack(fill="x", padx=6, pady=(5, 1))

        # Barra de progreso
        bar_c = tk.Canvas(card, bg=BG4, height=5, highlightthickness=0)
        bar_c.pack(fill="x", padx=6, pady=2)
        fill = bar_c.create_rectangle(0, 0, 0, 5, fill=ORANGE, outline="")

        def _on_bar_resize(e, bc=bar_c, f=fill, d={"pct": 0}):
            bc.coords(f, 0, 0, int(bc.winfo_width() * d["pct"]), 5)
        bar_c.bind("<Configure>", _on_bar_resize)

        # Estado + contador
        bot = tk.Frame(card, bg=BG3)
        bot.pack(fill="x", padx=6, pady=(0, 5))
        lbl_estado = tk.Label(bot, text="Abriendo...", bg=BG3,
                               fg=WHITE4, font=("Segoe UI", 7))
        lbl_estado.pack(side="left")
        lbl_cnt = tk.Label(bot, text="", bg=BG3,
                            fg=WHITE4, font=("Segoe UI", 7))
        lbl_cnt.pack(side="right")

        self._tarjetas_tienda[nombre] = {
            "frame":      card,
            "canvas":     bar_c,
            "fill":       fill,
            "bar_data":   {"pct": 0},
            "lbl_estado": lbl_estado,
            "lbl_cnt":    lbl_cnt,
            "listo":      0,
            "falta":      0,
            "dup":        0,
            "total":      0,
        }

        # Scroll al fondo
        self._tiendas_canvas.update_idletasks()
        self._tiendas_canvas.yview_moveto(1.0)
        self._actualizar_cnt_tiendas()

    def _tienda_card_update(self, nombre, resultado, motivo_falta=""):
        """
        Actualiza la tarjeta de una tienda con el resultado de una fila.
        resultado: 'LISTO' | 'FALTA' | 'DUP'
        motivo_falta: texto descriptivo del motivo si es FALTA
        """
        if nombre not in self._tarjetas_tienda:
            self._tienda_card_crear(nombre)

        t = self._tarjetas_tienda[nombre]
        t["total"] += 1

        if resultado == "LISTO":
            t["listo"] += 1
        elif resultado == "FALTA":
            t["falta"] += 1
        elif resultado == "DUP":
            t["dup"] += 1

        # Calcular porcentaje — avanza proporcionalmente
        total = t["total"]
        listo = t["listo"]
        pct   = min(1.0, listo / max(total, 1)) if total > 0 else 0

        # Actualizar barra
        t["bar_data"]["pct"] = pct
        w = t["canvas"].winfo_width()
        t["canvas"].coords(t["fill"], 0, 0, int(w * pct), 5)

        # Actualizar contador
        partes = []
        if t["listo"]: partes.append("✓{}".format(t["listo"]))
        if t["falta"]: partes.append("✗{}".format(t["falta"]))
        if t["dup"]:   partes.append("⚠{}".format(t["dup"]))
        t["lbl_cnt"].configure(text=" ".join(partes))

        # Estado textual
        if resultado == "LISTO":
            t["lbl_estado"].configure(text="Insertando...", fg=GREEN)
        elif resultado == "FALTA":
            motivo = motivo_falta or "Tienda no encontrada"
            t["lbl_estado"].configure(text="✗ " + motivo[:30], fg=RED)
        elif resultado == "DUP":
            t["lbl_estado"].configure(text="⚠ Duplicado", fg=YELLOW)

    def _tienda_card_cerrar(self, nombre, ok=True):
        """
        Marca la tarjeta de una tienda como finalizada.
        Cambia el borde: verde=ok, rojo=solo faltas, naranja=tiene dups.
        """
        if nombre not in self._tarjetas_tienda:
            return
        t = self._tarjetas_tienda[nombre]

        # Completar barra al 100%
        t["bar_data"]["pct"] = 1.0
        w = t["canvas"].winfo_width()
        t["canvas"].coords(t["fill"], 0, 0, w, 5)

        # Determinar color del borde final
        if t["falta"] > 0 and t["listo"] == 0:
            border_color = RED
            color_barra  = RED
            estado_txt   = "✗ Sin ingresos"
            estado_fg    = RED
        elif t["falta"] > 0 or t["dup"] > 0:
            border_color = YELLOW
            color_barra  = YELLOW
            estado_txt   = "⚠ Parcial"
            estado_fg    = YELLOW
        else:
            border_color = GREEN
            color_barra  = GREEN
            estado_txt   = "✓ Completado"
            estado_fg    = GREEN

        t["frame"].configure(highlightbackground=border_color)
        t["canvas"].itemconfig(t["fill"], fill=color_barra)
        t["lbl_estado"].configure(text=estado_txt, fg=estado_fg)

    def _tienda_card_limpiar(self):
        """Limpia todas las tarjetas para un nuevo proceso."""
        for w in self._tiendas_frame.winfo_children():
            w.destroy()
        self._tarjetas_tienda.clear()
        self._actualizar_cnt_tiendas()

    def _actualizar_cnt_tiendas(self):
        n = len(self._tarjetas_tienda)
        self._lbl_tiendas_cnt.configure(
            text="{} tienda{}".format(n, "s" if n != 1 else "") if n else "")

    def _log_add(self, texto, tag=""):
        cm = {"ok":GREEN,"err":RED,"warn":YELLOW,
              "info":ORANGE3,"dim":WHITE4,"bold":WHITE,"":WHITE3}
        fg = cm.get(tag, WHITE3)
        ts = datetime.now().strftime("%H:%M:%S")

        # Traducir mensajes del backend a descripciones legibles
        ll = texto.lower().strip()
        texto_display = texto

        if "  listo  fila" in ll:
            partes = texto.split("fila", 1)
            resto  = partes[1].strip() if len(partes) > 1 else ""
            num    = resto.split(":")[0].strip()
            tienda = resto.split(":", 1)[-1].strip() if ":" in resto else resto
            texto_display = f"✓  Fila {num}  →  Insertado en: {tienda}"
        elif "  falta  fila" in ll:
            partes = texto.split("fila", 1)
            resto  = partes[1].strip() if len(partes) > 1 else ""
            num    = resto.split(":")[0].strip()
            tienda = resto.split(":", 1)[-1].strip() if ":" in resto else resto
            texto_display = f"✗  Fila {num}  →  Tienda no encontrada: {tienda}"
        elif "  dup    fila" in ll:
            partes = texto.split("fila", 1)
            resto  = partes[1].strip() if len(partes) > 1 else ""
            num    = resto.split(":")[0].strip()
            info   = resto.split(":", 1)[-1].strip() if ":" in resto else resto
            texto_display = f"⚠  Fila {num}  →  Duplicado — ID ya existe: {info}"
        elif ll.startswith("abriendo"):
            nombre = texto.split("Abriendo")[-1].strip().rstrip("...")
            texto_display = f"📂  Abriendo archivo de tienda: {nombre}"
        elif ll.startswith("guardando") and "ingreso" not in ll:
            nombre = texto.split("Guardando")[-1].strip().rstrip("...")
            texto_display = f"💾  Guardando cambios en: {nombre}"
        elif "[ok] guardado:" in ll:
            nombre = texto.split(":")[-1].strip()
            texto_display = f"✓  Archivo guardado correctamente: {nombre}"
        elif "guardando ingreso_masivo" in ll:
            texto_display = "💾  Guardando archivo principal INGRESO_MASIVO..."
        elif "[ok] ingreso_masivo guardado" in ll:
            texto_display = "✓  INGRESO_MASIVO guardado con todos los resultados"
        elif "cargando ingreso_masivo" in ll:
            texto_display = "📥  Cargando archivo maestro INGRESO_MASIVO..."
        elif "indexando carpeta" in ll:
            texto_display = "🔍  Escaneando carpeta de tiendas en disco..."
        elif ll.startswith("[ok]") and "archivos encontrados" in ll:
            n = texto.split("]")[-1].strip().split()[0]
            texto_display = f"✓  {n} archivos de tiendas encontrados"
        elif "procesando filas" in ll:
            texto_display = f"⚙  {texto.strip()}"
        elif "[lote]" in ll:
            texto_display = "💾  Guardando lote de archivos al disco (cada 10 tiendas)..."
        elif "hoja tiendas:" in ll:
            n = texto.split(":")[-1].strip().split()[0]
            texto_display = f"✓  Tabla de alias de tiendas cargada ({n} variantes)"
        elif "cache col e:" in ll:
            n = texto.split(":")[-1].strip().split()[0]
            texto_display = f"✓  Cache de índice cargado ({n} nombres)"
        elif "sin cache" in ll:
            texto_display = "⚠  Sin cache de índice — usa 'Indexar tiendas' primero"
        elif "completado" in ll and "listo:" in ll:
            texto_display = f"🏁  {texto.strip()}"
        elif ll.startswith("listo") and ":" in ll:
            texto_display = f"✓  Total procesados: {texto.split(':')[-1].strip()}"
        elif ll.startswith("falta") and ":" in ll:
            texto_display = f"✗  Total sin tienda: {texto.split(':')[-1].strip()}"
        elif ll.startswith("dup") and ":" in ll:
            texto_display = f"⚠  Total duplicados: {texto.split(':')[-1].strip()}"
        elif "tiempo:" in ll:
            texto_display = f"⏱  {texto.strip()}"
        elif "[error]" in ll:
            texto_display = f"✗  Error: {texto.split(']',1)[-1].strip()}"
        elif "actualizando hoja falta" in ll:
            texto_display = "📋  Escribiendo registros FALTA en hoja de reporte..."
        elif "filas en hoja falta" in ll:
            n = texto.split("[ok]")[-1].strip().split()[0]
            texto_display = f"✓  {n} registros escritos en hoja FALTA"

        # Limitar a 600 widgets
        hijos = self._lframe.winfo_children()
        if len(hijos) > 600:
            for w in hijos[:100]:
                w.destroy()

        row = tk.Frame(self._lframe, bg=BG3)
        row.pack(fill="x")

        # Borde izquierdo segun tipo
        border_col = {"ok":GREEN,"err":RED,"warn":YELLOW,
                      "info":ORANGE,"bold":ORANGE}.get(tag, BORDER)
        tk.Frame(row, bg=border_col, width=2).pack(side="left", fill="y")

        tk.Label(row, text=ts, bg=BG3, fg=WHITE4,
                 font=("Consolas",8), width=9).pack(side="left", padx=(6,4))
        tk.Label(row, text=texto_display, bg=BG3, fg=fg,
                 font=("Segoe UI",9), anchor="w",
                 wraplength=380, justify="left").pack(
                     side="left", fill="x", expand=True,
                     padx=(0,10), pady=3)
        tk.Frame(self._lframe, bg=BORDER, height=1).pack(fill="x")

        # Actualizar scroll
        self._lcanvas.after_idle(lambda: (
            self._lcanvas.configure(
                scrollregion=self._lcanvas.bbox("all")),
            self._lcanvas.yview_moveto(1.0)
        ))

        # Actualizar estado detallado
        if hasattr(self, '_lbl_estado_det'):
            self._lbl_estado_det.configure(text=texto_display[:80], fg=fg)

    def _limpiar_log(self):
        for w in self._lframe.winfo_children(): w.destroy()

    def _tick_fecha(self):
        self._lbl_fecha.configure(text=fmt_hora())
        self.after(30000, self._tick_fecha)

    def _tick_footer(self):
        self._lbl_footer_hora.configure(
            text=datetime.now().strftime("%I:%M %p"))
        self.after(5000, self._tick_footer)

    def _refrescar_stats(self):
        s = cargar_stats()
        self._set_card(self._c_listo,  s["listo"])
        self._set_card(self._c_falta,  s["falta"])
        self._set_card(self._c_dup,    s["dup"])
        self._set_card(self._c_tiempo, f"{s['tiempo']}s" if s["tiempo"] else "--")
        self._set_card(self._c_fecha,  s["fecha"] or "--")
        self._build_resumen()

    def _set_prog_real(self, pct, proc, total):
        # Animacion suave: avanzar de a poco hacia el target
        current = self._prog_real["value"]
        target  = float(pct)
        if target > current:
            step = max(0.5, (target - current) * 0.3)
            nuevo = min(target, current + step)
            self._prog_real["value"] = nuevo
            if nuevo < target:
                self.after(30, self._set_prog_real, pct, proc, total)
                return
        self._prog_real["value"] = target
        self._lbl_pct_real.configure(text=f"{int(target)}%")
        self._lbl_filas.configure(text=f"{proc} / {total} filas")

        # Calcular ETA basado en velocidad real
        if proc > 2 and total > 0 and self._t_inicio:
            elapsed = time.time() - self._t_inicio
            vel     = proc / elapsed          # filas/segundo
            restantes = total - proc
            if vel > 0:
                eta_s = restantes / vel
                if eta_s < 60:
                    eta_txt = f"ETA: {int(eta_s)}s"
                else:
                    eta_txt = f"ETA: {int(eta_s/60)}m {int(eta_s%60)}s"
                self._lbl_eta.configure(text=eta_txt, fg=GREEN if eta_s < 30 else BLUE)
            else:
                self._lbl_eta.configure(text="ETA: --", fg=WHITE4)
        elif proc == 0:
            self._lbl_eta.configure(text="ETA: --", fg=WHITE4)

    def _set_actividad(self, txt, color=WHITE3):
        # Mapear texto a icono y descripcion detallada
        ll = txt.lower()
        if "listo" in ll and "fila" in ll:
            ico = "✓"; color = GREEN
            partes = txt.split("fila")
            fila = partes[1].strip().split(":")[0].strip() if len(partes) > 1 else ""
            tienda = partes[1].split(":")[-1].strip() if ":" in partes[-1] else ""
            txt = f"Fila {fila} procesada → {tienda}"
        elif "falta" in ll and "fila" in ll:
            ico = "✗"; color = YELLOW
            partes = txt.split("fila")
            fila = partes[1].strip().split(":")[0].strip() if len(partes) > 1 else ""
            tienda = partes[1].split(":")[-1].strip() if ":" in partes[-1] else txt
            txt = f"Fila {fila} sin tienda → {tienda}"
        elif "dup" in ll and "fila" in ll:
            ico = "⚠"; color = YELLOW
            txt = f"Duplicado detectado — {txt.split('fila')[-1].strip()}"
        elif "abriendo" in ll:
            ico = "📂"; color = ORANGE3
            txt = txt.replace("Abriendo","Abriendo archivo:").replace("abriendo","Abriendo archivo:")
        elif "guardando" in ll:
            ico = "💾"; color = BLUE
        elif "cargando" in ll:
            ico = "📥"; color = ORANGE3
        elif "indexando" in ll:
            ico = "🔍"; color = ORANGE3
        elif "procesando" in ll:
            ico = "⚙"; color = ORANGE
        elif "completado" in ll or "completo" in ll:
            ico = "🏁"; color = GREEN
        elif "error" in ll:
            ico = "✗"; color = RED
        elif "lote" in ll:
            ico = "💾"; color = BLUE
            txt = f"Guardando lote de archivos al disco..."
        else:
            ico = "▶"

        self._lbl_fase_ico.configure(text=ico, fg=color)
        self._lbl_actividad.configure(text=txt[:100], fg=color)

        # Elapsed timer
        if self._t_inicio:
            el = time.time() - self._t_inicio
            self._lbl_elapsed.configure(
                text=f"{int(el//60):02d}:{int(el%60):02d} transcurrido")

    def _pulse_dot(self, color=ORANGE, step=0):
        if not self._proceso_activo:
            self._dot_cv.itemconfig(self._dot, fill=WHITE4)
            return
        colors = [ORANGE, ORANGE2, ORANGE3, ORANGE2]
        self._dot_cv.itemconfig(self._dot, fill=colors[step % len(colors)])
        self._pulse_after = self.after(400, self._pulse_dot, color, step+1)

    def _bloquear(self):
        self._proceso_activo = True
        self._tiempos_fila   = []
        self._lbl_sb_estado.configure(text="● Procesando...", fg=ORANGE)
        self._lbl_eta.configure(text="ETA: calculando...", fg=WHITE4)
        self._lbl_elapsed.configure(text="00:00 transcurrido")
        self._lbl_fase_ico.configure(text="⚙", fg=ORANGE)
        self._lbl_actividad.configure(text="Iniciando proceso...", fg=ORANGE)
        self._prog_act.start(20)
        self._pulse_dot()
        self._tick_elapsed()

    def _desbloquear(self):
        self._proceso_activo = False
        self._lbl_sb_estado.configure(text="● Listo", fg=GREEN)
        self._lbl_eta.configure(text="ETA: --", fg=WHITE4)
        self._lbl_fase_ico.configure(text="⏸", fg=WHITE4)
        self._prog_act.stop()
        self._prog_act["value"] = 0
        if self._pulse_after:
            self.after_cancel(self._pulse_after)
        if self._eta_after:
            self.after_cancel(self._eta_after)
        self._dot_cv.itemconfig(self._dot, fill=WHITE4)

    def _tick_elapsed(self):
        """Actualiza el contador de tiempo transcurrido cada segundo."""
        if not self._proceso_activo:
            return
        if self._t_inicio:
            el = time.time() - self._t_inicio
            self._lbl_elapsed.configure(
                text=f"{int(el//60):02d}:{int(el%60):02d} transcurrido")
        self._eta_after = self.after(1000, self._tick_elapsed)

    # ── DIALOGO CAMBIAR RUTAS ────────────────────────────────────────────────
    def _abrir_config(self):
        self._pedir_pass("7070", self._do_cambiar_rutas)

    def _pedir_pass(self, clave, callback):
        dlg = tk.Toplevel(self)
        dlg.title("Autenticacion requerida")
        dlg.geometry("360x200")
        dlg.configure(bg=BG2)
        dlg.resizable(False,False)
        dlg.grab_set()
        dlg.transient(self)

        tk.Frame(dlg, bg=ORANGE, height=3).pack(fill="x")

        tk.Label(dlg, text="Acceso restringido", bg=BG2, fg=WHITE,
                 font=("Segoe UI",12,"bold")).pack(pady=(18,4))
        tk.Label(dlg, text="Ingresa la contrasena de administrador",
                 bg=BG2, fg=WHITE3, font=UI_SM).pack()

        ent = tk.Entry(dlg, show="*", bg=BG4, fg=WHITE, font=UI,
                       insertbackground=ORANGE, relief="flat", bd=0,
                       highlightbackground=BORDER2,
                       highlightthickness=1, width=26)
        ent.pack(pady=14, ipady=7)
        ent.focus_set()

        err_lbl = tk.Label(dlg, text="", bg=BG2, fg=RED, font=UI_SM)
        err_lbl.pack()

        def _ok(e=None):
            if ent.get() == clave:
                dlg.destroy(); callback()
            else:
                ent.configure(highlightbackground=RED)
                ent.delete(0,"end")
                err_lbl.configure(text="Contrasena incorrecta")

        ent.bind("<Return>", _ok)
        bf = tk.Frame(dlg, bg=BG2)
        bf.pack(pady=(4,0))
        tk.Button(bf, text="Cancelar", bg=BG3, fg=WHITE3,
                  font=UI_SM, relief="flat", cursor="hand2",
                  bd=0, padx=14, pady=6, activebackground=BG4,
                  command=dlg.destroy).pack(side="left", padx=4)
        tk.Button(bf, text="Ingresar", bg=ORANGE, fg=BG,
                  font=UI_B, relief="flat", cursor="hand2",
                  bd=0, padx=18, pady=6, activebackground=ORANGE2,
                  command=_ok).pack(side="left", padx=4)

    def _do_cambiar_rutas(self):
        # Leer config actual
        cfg_path = CONFIG_PY
        try:
            import importlib.util
            spec = importlib.util.spec_from_file_location("config_local", cfg_path)
            cfg  = importlib.util.module_from_spec(spec)
            spec.loader.exec_module(cfg)
            carpeta_actual = getattr(cfg, "CARPETA_TIENDAS", "")
            ingreso_actual = getattr(cfg, "ARCHIVO_INGRESO", "")
        except:
            carpeta_actual = ""
            ingreso_actual = ""

        dlg = tk.Toplevel(self)
        dlg.title("Cambiar rutas")
        dlg.geometry("560x320")
        dlg.configure(bg=BG2)
        dlg.resizable(False,False)
        dlg.grab_set()
        dlg.transient(self)

        tk.Frame(dlg, bg=ORANGE, height=3).pack(fill="x")
        tk.Label(dlg, text="Configurar rutas", bg=BG2, fg=WHITE,
                 font=("Segoe UI",12,"bold")).pack(pady=(16,4))
        tk.Label(dlg, text="Selecciona las rutas de los archivos del sistema",
                 bg=BG2, fg=WHITE3, font=UI_SM).pack(pady=(0,12))

        def _fila_ruta(parent, label, valor_inicial):
            f = tk.Frame(parent, bg=BG2)
            f.pack(fill="x", padx=20, pady=4)
            tk.Label(f, text=label, bg=BG2, fg=WHITE3,
                     font=("Segoe UI",8,"bold"), width=20, anchor="w").pack(side="left")
            var = tk.StringVar(value=valor_inicial)
            ent = tk.Entry(f, textvariable=var, bg=BG4, fg=WHITE,
                           font=("Segoe UI",9), relief="flat", bd=0,
                           highlightbackground=BORDER2,
                           highlightthickness=1, width=34)
            ent.pack(side="left", ipady=5, padx=(0,4))
            return var, ent

        def _browse_folder(var):
            r = filedialog.askdirectory(title="Seleccionar carpeta de tiendas")
            if r: var.set(r)

        def _browse_file(var):
            r = filedialog.askopenfilename(
                title="Seleccionar INGRESO_MASIVO",
                filetypes=[("Excel", "*.xlsx *.xls"), ("Todos", "*.*")])
            if r: var.set(r)

        f1 = tk.Frame(dlg, bg=BG2)
        f1.pack(fill="x", padx=20, pady=4)
        tk.Label(f1, text="Carpeta tiendas:", bg=BG2, fg=WHITE3,
                 font=("Segoe UI",8,"bold"), width=20, anchor="w").pack(side="left")
        var_carpeta = tk.StringVar(value=carpeta_actual)
        ent1 = tk.Entry(f1, textvariable=var_carpeta, bg=BG4, fg=WHITE,
                        font=("Segoe UI",9), relief="flat", bd=0,
                        highlightbackground=BORDER2,
                        highlightthickness=1, width=28)
        ent1.pack(side="left", ipady=5, padx=(0,4))
        tk.Button(f1, text="Buscar", bg=BG5, fg=ORANGE,
                  font=("Segoe UI",8), relief="flat", cursor="hand2",
                  bd=0, padx=8, pady=4,
                  command=lambda: _browse_folder(var_carpeta)).pack(side="left")

        f2 = tk.Frame(dlg, bg=BG2)
        f2.pack(fill="x", padx=20, pady=4)
        tk.Label(f2, text="INGRESO_MASIVO:", bg=BG2, fg=WHITE3,
                 font=("Segoe UI",8,"bold"), width=20, anchor="w").pack(side="left")
        var_ingreso = tk.StringVar(value=ingreso_actual)
        ent2 = tk.Entry(f2, textvariable=var_ingreso, bg=BG4, fg=WHITE,
                        font=("Segoe UI",9), relief="flat", bd=0,
                        highlightbackground=BORDER2,
                        highlightthickness=1, width=28)
        ent2.pack(side="left", ipady=5, padx=(0,4))
        tk.Button(f2, text="Buscar", bg=BG5, fg=ORANGE,
                  font=("Segoe UI",8), relief="flat", cursor="hand2",
                  bd=0, padx=8, pady=4,
                  command=lambda: _browse_file(var_ingreso)).pack(side="left")

        def _guardar():
            nueva_carpeta = var_carpeta.get().strip()
            nuevo_ingreso = var_ingreso.get().strip()
            if not nueva_carpeta or not nuevo_ingreso:
                messagebox.showerror("Error", "Ambas rutas son requeridas", parent=dlg)
                return
            try:
                with open(cfg_path, "r", encoding="utf-8") as f:
                    texto = f.read()
                import re
                texto = re.sub(
                    r'CARPETA_TIENDAS\s*=\s*r?"[^"]*"',
                    f'CARPETA_TIENDAS = r"{nueva_carpeta}"', texto)
                texto = re.sub(
                    r'ARCHIVO_INGRESO\s*=\s*r?"[^"]*"',
                    f'ARCHIVO_INGRESO = r"{nuevo_ingreso}"', texto)
                with open(cfg_path, "w", encoding="utf-8") as f:
                    f.write(texto)
                dlg.destroy()
                self._log_add("Rutas actualizadas correctamente.", "ok")
                messagebox.showinfo("Guardado",
                    "Rutas actualizadas.\nReinicia el programa para aplicar los cambios.",
                    parent=self)
            except Exception as e:
                messagebox.showerror("Error", str(e), parent=dlg)

        bf = tk.Frame(dlg, bg=BG2)
        bf.pack(pady=16)
        tk.Button(bf, text="Cancelar", bg=BG3, fg=WHITE3,
                  font=UI_SM, relief="flat", cursor="hand2",
                  bd=0, padx=14, pady=6, activebackground=BG4,
                  command=dlg.destroy).pack(side="left", padx=4)
        tk.Button(bf, text="Guardar rutas", bg=ORANGE, fg=BG,
                  font=UI_B, relief="flat", cursor="hand2",
                  bd=0, padx=18, pady=6, activebackground=ORANGE2,
                  command=_guardar).pack(side="left", padx=4)

    # ── NOTIFICACION WINDOWS ─────────────────────────────────────────────────
    def _notificar(self, titulo, mensaje):
        try:
            from plyer import notification
            notification.notify(title=titulo, message=mensaje,
                                app_name="Xpress Ingreso Masivo",
                                timeout=8)
        except:
            try:
                subprocess.Popen([
                    "powershell", "-WindowStyle", "Hidden", "-Command",
                    f'[Windows.UI.Notifications.ToastNotificationManager, Windows.UI.Notifications, ContentType=WindowsRuntime] | Out-Null;'
                    f'$t = [Windows.UI.Notifications.ToastNotificationManager]::GetTemplateContent([Windows.UI.Notifications.ToastTemplateType]::ToastText02);'
                    f'$t.GetElementsByTagName("text")[0].AppendChild($t.CreateTextNode("{titulo}")) | Out-Null;'
                    f'$t.GetElementsByTagName("text")[1].AppendChild($t.CreateTextNode("{mensaje}")) | Out-Null;'
                    f'$n = [Windows.UI.Notifications.ToastNotification]::new($t);'
                    f'[Windows.UI.Notifications.ToastNotificationManager]::CreateToastNotifier("Xpress").Show($n);'
                ], creationflags=0x08000000)
            except: pass

    # ── ACCIONES ─────────────────────────────────────────────────────────────
    def _ejecutar(self):
        if self._proceso_activo: return
        self._bloquear(); self._limpiar_log()
        self._prog_real["value"] = 0
        self._lbl_pct_real.configure(text="0%")
        self._lbl_filas.configure(text="0 / 0 filas")
        self._lbl_eta.configure(text="ETA: calculando...", fg=WHITE4)
        self._total_filas = 0; self._proc_filas = 0
        self._tiempos_fila = []
        self._log_add("▶  Iniciando proceso de ingreso masivo...", "info")
        self._t_inicio = time.time()
        threading.Thread(target=self._run, args=(MAIN_PY,), daemon=True).start()


    def _deshacer_proceso(self):
        """Revierte el último proceso restaurando los backups de los archivos de tienda."""
        if self._proceso_activo:
            self._log_add("⚠  No se puede deshacer mientras hay un proceso activo.", "warn")
            return

        # Importar main_local para verificar si hay algo que deshacer
        import sys, importlib
        sys.path.insert(0, BASE_DIR)
        try:
            import main_local as _ml
            importlib.reload(_ml)
            puede, info = _ml._deshacer.puede_deshacer()
        except Exception as e:
            self._log_add("✗  Error al verificar backup: {}".format(e), "error")
            return

        if not puede:
            self._log_add("⚠  {}".format(info), "warn")
            return

        # Pedir contraseña antes de continuar
        import tkinter as _tk
        from tkinter import messagebox

        dlg_pass = _tk.Toplevel(self)
        dlg_pass.title("Contraseña requerida")
        dlg_pass.configure(bg=BG2)
        dlg_pass.resizable(False, False)
        dlg_pass.transient(self)
        dlg_pass.grab_set()

        # Centrar sobre la ventana principal
        self.update_idletasks()
        x = self.winfo_x() + (self.winfo_width()  // 2) - 175
        y = self.winfo_y() + (self.winfo_height() // 2) - 80
        dlg_pass.geometry("350x160+{}+{}".format(x, y))

        _tk.Frame(dlg_pass, bg=RED, height=3).pack(fill="x")

        body = _tk.Frame(dlg_pass, bg=BG2, padx=24, pady=18)
        body.pack(fill="both", expand=True)

        _tk.Label(body, text="Ingresa la contraseña para deshacer:",
                  bg=BG2, fg=WHITE3, font=("Segoe UI", 9)).pack(anchor="w")

        entry_pass = _tk.Entry(body, show="*", bg=BG4, fg=WHITE,
                               font=("Segoe UI", 10), relief="flat",
                               insertbackground=WHITE, bd=0)
        entry_pass.pack(fill="x", pady=(6, 12), ipady=5)
        entry_pass.focus_set()

        _pass_ok = [False]

        def _verificar(event=None):
            if entry_pass.get() == "IngresoMas2026x":
                _pass_ok[0] = True
                dlg_pass.destroy()
            else:
                entry_pass.delete(0, "end")
                entry_pass.configure(bg="#4a1a1a")
                dlg_pass.after(600, lambda: entry_pass.configure(bg=BG4))

        def _cancelar():
            dlg_pass.destroy()

        btn_frame = _tk.Frame(body, bg=BG2)
        btn_frame.pack(fill="x")
        _tk.Button(btn_frame, text="Cancelar", bg=BG3, fg=WHITE3,
                   font=("Segoe UI", 8), relief="flat", cursor="hand2",
                   bd=0, padx=12, pady=4, activebackground=BG5,
                   command=_cancelar).pack(side="left")
        _tk.Button(btn_frame, text="  Confirmar  ", bg=RED, fg=WHITE,
                   font=("Segoe UI", 8, "bold"), relief="flat", cursor="hand2",
                   bd=0, padx=12, pady=4, activebackground="#cc3355",
                   command=_verificar).pack(side="right")

        entry_pass.bind("<Return>", _verificar)
        entry_pass.bind("<Escape>", lambda e: _cancelar())

        self.wait_window(dlg_pass)

        if not _pass_ok[0]:
            return

        # Confirmar con el usuario
        conf = messagebox.askyesno(
            "Deshacer proceso",
            "{}\n\n¿Deseas restaurar los archivos de tienda a como estaban ANTES del último proceso?\n\nEsta acción no se puede deshacer.",
            icon="warning"
        )
        if not conf:
            return

        self._bloquear()
        self._limpiar_log()
        self._log_add("⟲  Deshaciendo último proceso de ingreso...", "info")

        def _run_deshacer():
            try:
                exito, msg = _ml.deshacer_ultimo_proceso(
                    callback_log=lambda m: self.after(0, lambda: self._log_add(m, "ok" if "[OK]" in m else "warn"))
                )
                def _fin():
                    if exito:
                        self._log_add("✓  {}".format(msg), "ok")
                    else:
                        self._log_add("✗  {}".format(msg), "error")
                    self._desbloquear()
                self.after(0, _fin)
            except Exception as e:
                self.after(0, lambda: self._log_add("✗  Error inesperado: {}".format(e), "error"))
                self.after(0, self._desbloquear)

        threading.Thread(target=_run_deshacer, daemon=True).start()

    def _verificar(self):
        if self._proceso_activo: return
        self._bloquear(); self._limpiar_log()
        self._log_add("🔍  Verificando integridad de archivos .xlsx...", "info")
        self._t_inicio = time.time()
        threading.Thread(target=self._run, args=(TEST_PY,), daemon=True).start()

    def _ver_falta(self):
        """Abre ventana dedicada con todos los paquetes no ingresados."""
        dlg = tk.Toplevel(self)
        dlg.title("Paquetes No Ingresados")
        dlg.geometry("860x560")
        dlg.configure(bg=BG2)
        dlg.resizable(True, True)
        dlg.transient(self)

        tk.Frame(dlg, bg=YELLOW, height=3).pack(fill="x")

        # Header
        hdr = tk.Frame(dlg, bg=BG2, padx=20, pady=12)
        hdr.pack(fill="x")
        tk.Label(hdr, text="Paquetes No Ingresados", bg=BG2, fg=WHITE,
                 font=("Segoe UI", 13, "bold")).pack(side="left")
        self._falta_lbl_total = tk.Label(hdr, text="Buscando...", bg=BG2,
                                          fg=YELLOW, font=("Segoe UI", 9, "bold"))
        self._falta_lbl_total.pack(side="right")

        # Columnas header
        cols_f = tk.Frame(dlg, bg=BG4, padx=16, pady=6)
        cols_f.pack(fill="x", padx=16)
        tk.Label(cols_f, text="FILA",   bg=BG4, fg=WHITE4, font=("Segoe UI",8,"bold"), width=5,  anchor="w").pack(side="left")
        tk.Label(cols_f, text="TIENDA", bg=BG4, fg=WHITE4, font=("Segoe UI",8,"bold"), width=30, anchor="w").pack(side="left", padx=(8,0))
        tk.Label(cols_f, text="ID",     bg=BG4, fg=WHITE4, font=("Segoe UI",8,"bold"), width=12, anchor="w").pack(side="left", padx=(8,0))
        tk.Label(cols_f, text="MOTIVO", bg=BG4, fg=WHITE4, font=("Segoe UI",8,"bold"), width=20, anchor="w").pack(side="left", padx=(8,0))

        tk.Frame(dlg, bg=BORDER2, height=1).pack(fill="x", padx=16)

        # Lista scrollable
        list_outer = tk.Frame(dlg, bg=BG2)
        list_outer.pack(fill="both", expand=True, padx=16, pady=(4,0))

        canvas = tk.Canvas(list_outer, bg=BG3, highlightthickness=0)
        sb_r   = tk.Scrollbar(list_outer, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=sb_r.set)
        sb_r.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)

        lista_frame = tk.Frame(canvas, bg=BG3)
        win_id = canvas.create_window((0,0), window=lista_frame, anchor="nw")
        lista_frame.bind("<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.bind("<Configure>",
            lambda e: canvas.itemconfig(win_id, width=e.width))
        canvas.bind("<MouseWheel>",
            lambda e: canvas.yview_scroll(int(-1*(e.delta/120)), "units"))

        # Footer con botón actualizar
        foot = tk.Frame(dlg, bg=BG2, padx=16, pady=10)
        foot.pack(fill="x")
        tk.Button(foot, text="Actualizar", bg=BG4, fg=WHITE3,
                  font=("Segoe UI",8), relief="flat", cursor="hand2",
                  bd=0, padx=12, pady=5, activebackground=BG5,
                  command=lambda: self._cargar_falta_ventana(
                      lista_frame, self._falta_lbl_total)).pack(side="left")
        tk.Label(foot, text="Muestra los registros con resultado FALTA o DUP del ultimo proceso",
                 bg=BG2, fg=WHITE4, font=("Segoe UI",7)).pack(side="left", padx=12)

        # Cargar datos
        self._cargar_falta_ventana(lista_frame, self._falta_lbl_total)

    def _cargar_falta_ventana(self, frame, lbl_total):
        """Carga los registros FALTA y DUP en la ventana dedicada."""
        for w in frame.winfo_children():
            w.destroy()
        tk.Label(frame, text="Cargando...", bg=BG3, fg=ORANGE,
                 font=("Segoe UI",9)).pack(padx=16, pady=12)
        lbl_total.configure(text="Buscando...")
        threading.Thread(
            target=self._leer_falta_ventana,
            args=(frame, lbl_total), daemon=True).start()

    def _indexar(self):
        if self._proceso_activo: return
        self._bloquear(); self._limpiar_log()
        self._log_add("🔍  Iniciando escaneo de columna E en todas las tiendas...", "info")
        self._t_inicio = time.time()
        threading.Thread(target=self._run_indexar, daemon=True).start()

    # ── RUNNER PRINCIPAL ─────────────────────────────────────────────────────
    def _run(self, script):
        import queue as _queue
        self._log_queue = _queue.Queue()
        c_listo = c_falta = c_dup = 0

        # Limpiar tarjetas del proceso anterior
        self.after(0, self._tienda_card_limpiar)

        def _procesar_cola():
            """Drena la cola de mensajes de a poco, suavemente."""
            procesados = 0
            while not self._log_queue.empty() and procesados < 8:
                try:
                    tipo, datos = self._log_queue.get_nowait()
                    if tipo == "log":
                        self._log_add(*datos)
                    elif tipo == "prog":
                        self._set_prog_real(*datos)
                    elif tipo == "act":
                        self._set_actividad(*datos)
                    elif tipo == "tarjeta":
                        nombre_t, resultado, motivo = datos
                        self._tienda_card_update(nombre_t, resultado, motivo)
                    elif tipo == "tarjeta_crear":
                        self._tienda_card_crear(datos[0])
                    elif tipo == "tarjeta_cerrar":
                        # Buscar la tarjeta por nombre parcial
                        nombre_arch = datos[0]
                        for k in list(self._tarjetas_tienda.keys()):
                            if nombre_arch.lower() in k.lower() or k.lower() in nombre_arch.lower():
                                self._tienda_card_cerrar(k)
                                break
                    procesados += 1
                except: break
            if self._proceso_activo or not self._log_queue.empty():
                self.after(60, _procesar_cola)

        self.after(60, _procesar_cola)

        try:
            proc = subprocess.Popen(
                [sys.executable, script],
                stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
                text=True, encoding="utf-8", errors="replace", cwd=BASE_DIR)

            # Fases del proceso con porcentajes aproximados
            # para que la barra se mueva desde el inicio
            FASES = [
                ("cargando",    3),   # Cargando INGRESO_MASIVO
                ("indexando",   6),   # Indexando carpeta
                ("cache",       8),   # Cache col E
                ("procesando", 10),   # Empieza el loop — aquí tomamos control real
            ]
            fase_pct = 0

            for linea in proc.stdout:
                linea = linea.rstrip()
                if not linea: continue
                ll = linea.lower()
                ls = ll.strip()

                # Detectar total de filas — formato confiable TOTAL_FILAS:N
                if ll.startswith("total_filas:"):
                    try:
                        self._total_filas = int(ll.split(":")[1].strip())
                    except: pass

                # Fallback: parseo de la linea "Procesando filas X a Y"
                elif "procesando" in ll and "tiendas unicas" in ll:
                    try:
                        import re as _re
                        m = _re.search(r'(\d+)\s+tiendas', ll)
                        m2 = _re.search(r'filas\s+\d+\s+a\s+(\d+)', ll)
                        m3 = _re.search(r'filas\s+(\d+)\s+a', ll)
                        if m2 and m3:
                            self._total_filas = int(m2.group(1)) - int(m3.group(1)) + 1
                    except: pass

                # Avanzar barra en fases iniciales antes de tener el total
                if self._total_filas == 0:
                    for kw, pct in FASES:
                        if kw in ll and pct > fase_pct:
                            fase_pct = pct
                            self._log_queue.put(("prog", (fase_pct, 0, 0)))
                            break

                # Contar procesadas — progreso real fila a fila
                if "  listo  fila" in ll or "  falta  fila" in ll or "  dup    fila" in ll:
                    self._proc_filas += 1
                    self._tiempos_fila.append(time.time())
                    if len(self._tiempos_fila) > 20:
                        self._tiempos_fila = self._tiempos_fila[-20:]
                    if self._total_filas > 0:
                        pct = 10 + min(85, self._proc_filas / self._total_filas * 85)
                        self._log_queue.put(("prog",
                            (pct, self._proc_filas, self._total_filas)))

                    # Extraer tienda y resultado para la tarjeta
                    try:
                        import re as _re2
                        if "  listo  fila" in ll:
                            m = _re2.search(r'listo\s+fila\s+\d+[:\s]+(.+)', linea, _re2.IGNORECASE)
                            nombre_t = m.group(1).strip() if m else "Tienda"
                            self._log_queue.put(("tarjeta", (nombre_t, "LISTO", "")))
                        elif "  dup    fila" in ll:
                            m = _re2.search(r'dup\s+fila\s+\d+[:\s]+(.+)', linea, _re2.IGNORECASE)
                            nombre_t = m.group(1).strip() if m else "Tienda"
                            self._log_queue.put(("tarjeta", (nombre_t, "DUP", "")))
                        elif "  falta  fila" in ll:
                            # Detectar motivo real del FALTA
                            motivo = "Tienda no encontrada"
                            if "hoja invalida" in ll:
                                motivo = "Archivo sin hoja válida"
                            elif "encabezados" in ll:
                                motivo = "Encabezados incorrectos"
                            elif "no se pudo abrir" in ll:
                                motivo = "No se pudo abrir el archivo"
                            elif "bloqueado" in ll:
                                motivo = "Archivo bloqueado por Excel"
                            else:
                                # Extraer nombre de la tienda buscada
                                m = _re2.search(r"falta\s+fila\s+\d+[:\s]+'?([^']+)'?", linea, _re2.IGNORECASE)
                                nombre_t_raw = m.group(1).strip() if m else ""
                                if nombre_t_raw:
                                    motivo = "No encontrada: {}".format(nombre_t_raw[:25])
                            m = _re2.search(r'falta\s+fila\s+\d+[:\s]+(.+)', linea, _re2.IGNORECASE)
                            nombre_t = m.group(1).strip() if m else "Tienda"
                            # Limpiar comillas del nombre
                            nombre_t = nombre_t.strip("'\"")
                            self._log_queue.put(("tarjeta", (nombre_t, "FALTA", motivo)))
                    except Exception:
                        pass

                # Detectar apertura de tienda
                if ll.startswith("  abriendo "):
                    import re as _re3
                    m = _re3.search(r'abriendo\s+(.+?)\.\.\.', linea, _re3.IGNORECASE)
                    if m:
                        nombre_t = m.group(1).strip()
                        self._log_queue.put(("tarjeta_crear", (nombre_t,)))

                # Detectar guardado de tienda (cierre de tarjeta)
                if "[ok] guardado:" in ll:
                    import re as _re4
                    m = _re4.search(r'guardado:\s*(.+)', linea, _re3.IGNORECASE)
                    if m:
                        nombre_arch = m.group(1).strip()
                        # Quitar extension
                        nombre_sin_ext = nombre_arch.rsplit(".", 1)[0] if "." in nombre_arch else nombre_arch
                        self._log_queue.put(("tarjeta_cerrar", (nombre_sin_ext,)))
                if ls.startswith("listo") and ":" in ls:
                    try: c_listo = int(linea.split(":")[-1].strip().split()[0])
                    except: pass
                elif ls.startswith("falta") and ":" in ls:
                    try: c_falta = int(linea.split(":")[-1].strip().split()[0])
                    except: pass
                elif ls.startswith("dup") and ":" in ls:
                    try: c_dup = int(linea.split(":")[-1].strip().split()[0])
                    except: pass

                # Tag de color
                if "  listo  fila" in ll:         tag = "ok"
                elif "  falta  fila" in ll:        tag = "warn"
                elif "  dup    fila" in ll:        tag = "warn"
                elif "[ok]" in ll:                 tag = "ok"
                elif "listo" in ls and ":" in ls:  tag = "ok"
                elif "falta" in ls and ":" in ls:  tag = "warn"
                elif "error" in ll:                tag = "err"
                elif any(x in ll for x in ["guardando","cargando","indexando",
                                            "leyendo","abriendo","procesando",
                                            "escaneando","agrupando"]): tag = "info"
                elif "===" in linea or "---" in linea: tag = "dim"
                else: tag = ""

                # Encolar log y actividad
                self._log_queue.put(("log", (linea.strip(), tag)))
                self._log_queue.put(("act", (linea.strip()[:80],)))

            proc.wait()
            el = time.time() - self._t_inicio if self._t_inicio else 0

            if proc.returncode == 0:
                self._log_queue.put(("prog", (100, self._total_filas, self._total_filas)))
                self._log_queue.put(("act",  ("Proceso completado",)))
                self._log_queue.put(("log",  (
                    "Completado en {:.1f}s  —  LISTO: {}  FALTA: {}  DUP: {}".format(
                        el, c_listo, c_falta, c_dup), "ok")))
                guardar_stats(c_listo, c_falta, c_dup, el)
                self.after(500, self._refrescar_stats)
                self.after(800, self._notificar,
                           "Xpress — Proceso completado",
                           "LISTO: {}  FALTA: {}  DUP: {}  ({:.1f}s)".format(
                               c_listo, c_falta, c_dup, el))
            else:
                self._log_queue.put(("act", ("Error en el proceso",)))
                self._log_queue.put(("log", ("El proceso termino con errores.", "err")))
        except Exception as e:
            self._log_queue.put(("log", (f"Excepcion: {e}", "err")))
        finally:
            self.after(0, self._desbloquear)

    def _run_indexar(self):
        try:
            import importlib.util
            spec = importlib.util.spec_from_file_location("indexar", INDEXAR_PY)
            idx  = importlib.util.module_from_spec(spec)
            spec.loader.exec_module(idx)
            spec2 = importlib.util.spec_from_file_location("config_local", CONFIG_PY)
            cfg  = importlib.util.module_from_spec(spec2)
            spec2.loader.exec_module(cfg)

            total_ref = [0]
            def _prog(actual, total):
                total_ref[0] = total
                pct = actual / max(total,1) * 100
                self.after(0, self._set_prog_real, pct, actual, total)
                self.after(0, self._set_actividad,
                           f"Indexando {actual}/{total} archivos...", ORANGE3)

            indice, n_arch, n_nombres = idx.indexar(cfg.CARPETA_TIENDAS, _prog)
            el = time.time() - self._t_inicio if self._t_inicio else 0
            self.after(0, self._set_prog_real, 100, n_arch, n_arch)
            self.after(0, self._log_add,
                       f"Indexado completo: {n_nombres} nombres en "
                       f"{n_arch} archivos ({el:.1f}s)", "ok")
            self.after(0, self._notificar, "Xpress — Indexado completo",
                       f"{n_nombres} nombres indexados en {n_arch} archivos")
        except Exception as e:
            self.after(0, self._log_add, f"Error al indexar: {e}", "err")
        finally:
            self.after(0, self._desbloquear)

    def _leer_falta_ventana(self, frame, lbl_total):
        """Lee FALTA y DUP del INGRESO_MASIVO y los muestra en la ventana."""
        try:
            import importlib.util, zipfile, re as _re

            spec = importlib.util.spec_from_file_location("config_local", CONFIG_PY)
            cfg  = importlib.util.module_from_spec(spec)
            spec.loader.exec_module(cfg)

            ruta = cfg.ARCHIVO_INGRESO

            if not os.path.isfile(ruta):
                self.after(0, lambda: self._falta_render_error(
                    frame, lbl_total, "No se encontró INGRESO_MASIVO:\n{}".format(ruta)))
                return

            _PAT_SS = _re.compile(rb'<si>.*?<t[^>]*>(.*?)</t>.*?</si>', _re.DOTALL)
            _PAT_C  = _re.compile(
                rb'<c r="([A-Z]+)(\d+)"[^>]*>'
                rb'(?:<is><t[^>]*>(.*?)</t></is>|<v>(.*?)</v>)',
                _re.DOTALL)

            with zipfile.ZipFile(ruta, 'r') as z:
                nz = z.namelist()
                strings = []
                for ss in ['xl/sharedStrings.xml', 'xl/SharedStrings.xml']:
                    if ss in nz:
                        strings = [m.group(1).decode('utf-8','ignore').strip()
                                   for m in _PAT_SS.finditer(z.read(ss))]
                        break
                hoja = None
                for cand in nz:
                    if 'worksheets/sheet' in cand and cand.endswith('.xml'):
                        hoja = cand; break
                if not hoja:
                    raise Exception("No se encontró la hoja en el archivo")
                raw = z.read(hoja)

            def _val(inline, v_raw):
                if inline:
                    return inline.decode('utf-8','ignore').strip()
                if v_raw:
                    s = v_raw.decode('utf-8','ignore').strip()
                    if s.isdigit() and strings:
                        idx = int(s)
                        return strings[idx] if idx < len(strings) else s
                    return s
                return ""

            # Leer cols B (resultado), E (tienda), F (ID), N (comentario)
            cols_leer = {'B', 'E', 'F', 'N'}
            filas = {}
            for m in _PAT_C.finditer(raw):
                col_l  = m.group(1).decode()
                fila_n = int(m.group(2))
                if fila_n < cfg.FILA_INICIO:
                    continue
                if col_l not in cols_leer:
                    continue
                v = _val(m.group(3), m.group(4))
                if fila_n not in filas:
                    filas[fila_n] = {}
                filas[fila_n][col_l] = v

            # Filtrar solo FALTA y DUP
            registros = []
            for fila_n in sorted(filas.keys()):
                datos = filas[fila_n]
                res   = datos.get('B', '').strip().upper()
                if res in ('FALTA', 'DUP'):
                    registros.append({
                        "fila":    fila_n,
                        "tienda":  datos.get('E', '--'),
                        "id":      datos.get('F', '--'),
                        "comentario": datos.get('N', ''),
                        "motivo":  res,
                    })

            self.after(0, self._falta_render, frame, lbl_total, registros)

        except Exception as e:
            self.after(0, lambda: self._falta_render_error(
                frame, lbl_total, "Error: {}".format(e)))

    def _falta_render_error(self, frame, lbl_total, msg):
        for w in frame.winfo_children(): w.destroy()
        tk.Label(frame, text=msg, bg=BG3, fg=RED,
                 font=("Segoe UI",9), wraplength=700,
                 justify="left").pack(padx=16, pady=16, anchor="w")
        lbl_total.configure(text="Error")

    def _falta_render(self, frame, lbl_total, registros):
        for w in frame.winfo_children(): w.destroy()

        n_falta = sum(1 for r in registros if r["motivo"] == "FALTA")
        n_dup   = sum(1 for r in registros if r["motivo"] == "DUP")

        if not registros:
            tk.Label(frame,
                     text="✓  No hay registros FALTA ni DUP — todos los paquetes fueron ingresados.",
                     bg=BG3, fg=GREEN, font=("Segoe UI",10,"bold"),
                     wraplength=700).pack(padx=16, pady=24)
            lbl_total.configure(text="Todo OK", fg=GREEN)
            return

        lbl_total.configure(
            text="FALTA: {}   DUP: {}   Total: {}".format(n_falta, n_dup, len(registros)),
            fg=YELLOW)

        for r in registros:
            motivo  = r["motivo"]
            color   = RED if motivo == "FALTA" else YELLOW
            bg_fila = BG4

            row_f = tk.Frame(frame, bg=bg_fila,
                             highlightbackground=color, highlightthickness=1)
            row_f.pack(fill="x", padx=8, pady=2)
            tk.Frame(row_f, bg=color, width=3).pack(side="left", fill="y")

            inner = tk.Frame(row_f, bg=bg_fila, padx=10, pady=6)
            inner.pack(side="left", fill="x", expand=True)

            top = tk.Frame(inner, bg=bg_fila)
            top.pack(fill="x")

            # Fila número
            tk.Label(top, text="#{:>5}".format(r["fila"]),
                     bg=bg_fila, fg=WHITE4, font=("Consolas",8), width=7,
                     anchor="w").pack(side="left")

            # Tienda
            tk.Label(top, text=r["tienda"], bg=bg_fila, fg=WHITE,
                     font=("Segoe UI",9,"bold"), width=30,
                     anchor="w").pack(side="left", padx=(6,0))

            # ID
            tk.Label(top, text="ID: {}".format(r["id"]), bg=bg_fila,
                     fg=WHITE3, font=("Segoe UI",8), width=14,
                     anchor="w").pack(side="left", padx=(6,0))

            # Motivo con descripción
            if motivo == "FALTA":
                desc = "Tienda no encontrada en ningún archivo"
            else:
                desc = "ID duplicado — ya existe en el archivo destino"

            tk.Label(top, text=desc, bg=bg_fila, fg=color,
                     font=("Segoe UI",8), anchor="w").pack(side="left", padx=(6,0))

            # Comentario si existe
            if r["comentario"]:
                tk.Label(inner, text="Comentario: {}".format(r["comentario"]),
                         bg=bg_fila, fg=WHITE4, font=("Segoe UI",7),
                         anchor="w").pack(anchor="w", pady=(2,0))

        # Separador final
        tk.Frame(frame, bg=BORDER, height=1).pack(fill="x", padx=8, pady=4)


    # ── DIAGNOSTICO DE TIENDA ────────────────────────────────────────────────
    def _abrir_diagnostico(self):
        dlg = tk.Toplevel(self)
        dlg.title("Diagnostico de Tienda")
        dlg.geometry("720x580")
        dlg.configure(bg=BG2)
        dlg.resizable(True, True)
        dlg.transient(self)

        tk.Frame(dlg, bg=RED, height=3).pack(fill="x")

        hdr = tk.Frame(dlg, bg=BG2, padx=20, pady=14)
        hdr.pack(fill="x")
        tk.Label(hdr, text="Diagnostico de Tienda", bg=BG2, fg=WHITE,
                 font=("Segoe UI", 13, "bold")).pack(anchor="w")
        tk.Label(hdr, text="Escribe el nombre exacto como aparece en INGRESO_MASIVO y presiona Diagnosticar",
                 bg=BG2, fg=WHITE3, font=("Segoe UI", 8)).pack(anchor="w", pady=(2, 0))

        search_f = tk.Frame(dlg, bg=BG2, padx=20)
        search_f.pack(fill="x", pady=(0, 10))

        self._diag_var = tk.StringVar()
        resultado_frame_ref = [None]

        ent = tk.Entry(search_f, textvariable=self._diag_var, bg=BG4, fg=WHITE,
                       font=("Segoe UI", 11), relief="flat", bd=0,
                       insertbackground=RED,
                       highlightbackground=BORDER2, highlightthickness=1)
        ent.pack(side="left", fill="x", expand=True, ipady=8, padx=(0, 8))
        ent.focus_set()

        def _lanzar():
            if resultado_frame_ref[0]:
                self._run_diagnostico(resultado_frame_ref[0])

        btn = tk.Button(search_f, text="Diagnosticar", bg=RED, fg=WHITE,
                        font=("Segoe UI", 9, "bold"), relief="flat", cursor="hand2",
                        bd=0, padx=16, pady=8, activebackground=ORANGE2,
                        command=_lanzar)
        btn.pack(side="left")
        ent.bind("<Return>", lambda e: _lanzar())

        tk.Frame(dlg, bg=BORDER2, height=1).pack(fill="x", padx=20, pady=(0, 6))

        res_outer = tk.Frame(dlg, bg=BG2)
        res_outer.pack(fill="both", expand=True, padx=20, pady=(0, 16))

        canvas = tk.Canvas(res_outer, bg=BG3, highlightthickness=0)
        sb_r   = tk.Scrollbar(res_outer, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=sb_r.set)
        sb_r.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)

        rf = tk.Frame(canvas, bg=BG3)
        win_id = canvas.create_window((0, 0), window=rf, anchor="nw")
        rf.bind("<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.bind("<Configure>",
            lambda e: canvas.itemconfig(win_id, width=e.width))
        canvas.bind("<MouseWheel>",
            lambda e: canvas.yview_scroll(int(-1*(e.delta/120)), "units"))

        resultado_frame_ref[0] = rf

        tk.Label(rf, text="Ingresa el nombre de una tienda para ver el diagnostico completo.",
                 bg=BG3, fg=WHITE4, font=("Segoe UI", 9),
                 wraplength=620, justify="left").pack(padx=16, pady=20, anchor="w")

    def _run_diagnostico(self, frame):
        nombre = self._diag_var.get().strip()
        if not nombre:
            return
        for w in frame.winfo_children():
            w.destroy()
        tk.Label(frame, text="Analizando...", bg=BG3, fg=ORANGE,
                 font=("Segoe UI", 9)).pack(padx=16, pady=12, anchor="w")
        threading.Thread(
            target=self._diagnostico_worker,
            args=(nombre, frame), daemon=True).start()

    def _diag_row(self, parent, icono, titulo, detalle, color):
        f = tk.Frame(parent, bg=BG4,
                     highlightbackground=color, highlightthickness=1)
        f.pack(fill="x", padx=12, pady=4)
        tk.Frame(f, bg=color, width=4).pack(side="left", fill="y")
        inner = tk.Frame(f, bg=BG4, padx=12, pady=10)
        inner.pack(side="left", fill="x", expand=True)
        top = tk.Frame(inner, bg=BG4)
        top.pack(fill="x", anchor="w")
        tk.Label(top, text=icono, bg=BG4, fg=color,
                 font=("Segoe UI", 12)).pack(side="left", padx=(0, 8))
        tk.Label(top, text=titulo, bg=BG4, fg=color,
                 font=("Segoe UI", 10, "bold"), wraplength=560,
                 justify="left").pack(side="left", anchor="w")
        if detalle:
            tk.Label(inner, text=detalle, bg=BG4, fg=WHITE3,
                     font=("Segoe UI", 8), wraplength=580,
                     justify="left").pack(anchor="w", pady=(4, 0))

    def _diagnostico_worker(self, nombre_raw, frame):
        resultados = []
        try:
            import importlib.util
            import glob as _glob
            import unicodedata as _uc
            import re as _re
            import json as _json
            import zipfile as _zf

            # Normalizar localmente — evita importar logica_local que tiene deps pesadas
            def _norm(txt):
                if txt is None:
                    return ""
                txt = str(txt).lower().strip()
                txt = _uc.normalize("NFD", txt)
                txt = "".join(c for c in txt if _uc.category(c) != "Mn")
                txt = _re.sub(r"[^a-z0-9]", "", txt)
                return txt

            # Cargar config directamente leyendo el archivo como texto
            # (evita ejecutar imports encadenados que pueden colgarse)
            cfg_path = CONFIG_PY
            carpeta       = ""
            archivo_ingreso = ""
            hoja_tiendas  = "TIENDAS"
            fila_tiendas_ini = 2
            fila_encabezado  = 5
            fila_datos_dest  = 6
            encabezados_validos = {4:"F.RECOLECTA",5:"TIENDA",6:"ID",7:"NOMBRE",8:"ZONA",9:"TELEFONO",10:"PRECIO"}

            try:
                with open(cfg_path, "r", encoding="utf-8") as _f:
                    _txt = _f.read()
                for _line in _txt.splitlines():
                    _line = _line.strip()
                    if _line.startswith("#"):
                        continue
                    if "=" not in _line:
                        continue
                    _key, _, _val_raw = _line.partition("=")
                    _key = _key.strip()
                    _val_raw = _val_raw.strip().split("#")[0].strip()  # quitar comentarios inline
                    # Extraer valor entre comillas (maneja r"...", r'...', "...", '...')
                    _m = _re.search(r'[rR]?["\'](.+)["\']', _val_raw)
                    _val_str = _m.group(1) if _m else _val_raw.strip("\"'rR ")
                    if _key == "CARPETA_TIENDAS":
                        carpeta = _val_str
                    elif _key == "ARCHIVO_INGRESO":
                        archivo_ingreso = _val_str
                    elif _key == "HOJA_TIENDAS":
                        hoja_tiendas = _val_str
                    elif _key == "FILA_TIENDAS_INI":
                        try: fila_tiendas_ini = int(_val_str)
                        except: pass
                    elif _key == "FILA_ENCABEZADO":
                        try: fila_encabezado = int(_val_str)
                        except: pass
            except Exception as ex:
                resultados.append(("X", "No se pudo leer config_local.py", str(ex), RED))
                self.after(0, self._render_diagnostico, frame, resultados)
                return

            resultados = []
            norm = _norm(nombre_raw)

            resultados.append((">>", "Nombre analizado: \"{}\"".format(nombre_raw),
                "Normalizado internamente como: \"{}\"  (sin tildes, espacios ni caracteres especiales)".format(norm),
                WHITE3))

            # ── NIVEL 1: nombre exacto del archivo ────────────────────────
            if not os.path.isdir(carpeta):
                resultados.append(("X", "Carpeta de tiendas no encontrada",
                    "Ruta configurada: \"{}\"\nVerifica la ruta en 'Cambiar rutas'".format(carpeta), RED))
                self.after(0, self._render_diagnostico, frame, resultados)
                return

            archivos = {}
            for ruta in _glob.glob(os.path.join(carpeta, "*.xls*")):
                nb = os.path.basename(ruta)
                if nb.startswith("~$"):
                    continue
                dot     = nb.rfind(".")
                sin_ext = nb[:dot] if dot > 0 else nb
                k = _norm(sin_ext)
                if k:
                    archivos[k] = ruta

            resultados.append((">>", "{} archivos .xlsx encontrados en la carpeta".format(len(archivos)),
                "", WHITE4))

            if norm in archivos:
                resultados.append(("OK", "NIVEL 1 OK — Archivo encontrado por nombre exacto",
                    "Archivo: {}".format(os.path.basename(archivos[norm])), GREEN))
                self._diag_verificar_hoja_zip(archivos[norm], encabezados_validos,
                                               fila_encabezado, resultados)
                self.after(0, self._render_diagnostico, frame, resultados)
                return
            else:
                cercanos = self._diag_similares(norm, list(archivos.keys()), n=4)
                detalle = ("No hay archivo cuyo nombre normalizado sea \"{}\"\n"
                           "Archivos mas similares en la carpeta: {}").format(
                               norm, ", ".join(cercanos) if cercanos else "ninguno")
                resultados.append(("X", "NIVEL 1 FALLO — No encontrado por nombre de archivo",
                    detalle, RED))

            # ── NIVEL 2: hoja TIENDAS (leer ZIP directo, sin openpyxl) ────
            mapa_tiendas = {}
            if os.path.isfile(archivo_ingreso):
                try:
                    mapa_tiendas = self._diag_leer_hoja_tiendas_zip(
                        archivo_ingreso, hoja_tiendas, fila_tiendas_ini, _norm)
                except Exception as ex:
                    resultados.append(("!", "No se pudo leer la hoja TIENDAS",
                        str(ex), YELLOW))

            if norm in mapa_tiendas:
                destino = mapa_tiendas[norm]
                if destino in archivos:
                    resultados.append(("OK", "NIVEL 2 OK — Encontrado via hoja TIENDAS",
                        "Alias \"{}\" apunta al archivo \"{}\"".format(nombre_raw, destino), GREEN))
                    self._diag_verificar_hoja_zip(archivos[destino], encabezados_validos,
                                                   fila_encabezado, resultados)
                    self.after(0, self._render_diagnostico, frame, resultados)
                    return
                else:
                    resultados.append(("!", "NIVEL 2 FALLO — Alias en hoja TIENDAS pero archivo no existe en carpeta",
                        "El alias apunta a \"{}\" pero ese .xlsx no esta en ALL_TIENDAS".format(destino), YELLOW))
            else:
                if not mapa_tiendas:
                    resultados.append(("X", "NIVEL 2 FALLO — Hoja TIENDAS vacia o no encontrada en INGRESO_MASIVO",
                        "No se pudo leer ningun alias de la hoja TIENDAS", RED))
                else:
                    resultados.append(("X", "NIVEL 2 FALLO — Nombre no registrado en hoja TIENDAS",
                        "No hay alias para \"{}\" en la hoja TIENDAS ({} aliases registrados)".format(
                            nombre_raw, len(mapa_tiendas)), RED))

            # ── NIVEL 3: cache col E ───────────────────────────────────────
            cache_file = os.path.join(BASE_DIR, "cache_cole.json")
            indice_cole = {}
            if os.path.isfile(cache_file):
                try:
                    with open(cache_file, "r", encoding="utf-8") as _f:
                        indice_cole = _json.load(_f).get("indice", {})
                except Exception:
                    pass

            if not indice_cole:
                resultados.append(("!", "NIVEL 3 FALLO — Cache col E vacio o no existe",
                    "Ejecuta 'Indexar tiendas' para generarlo. Sin cache el nivel 3 no funciona.", YELLOW))
            elif norm in indice_cole:
                clave_dest = indice_cole[norm]
                if clave_dest in archivos:
                    resultados.append(("OK", "NIVEL 3 OK — Encontrado via cache col E",
                        "El nombre aparece en col E del archivo \"{}\"".format(clave_dest), GREEN))
                    self._diag_verificar_hoja_zip(archivos[clave_dest], encabezados_validos,
                                                   fila_encabezado, resultados)
                    self.after(0, self._render_diagnostico, frame, resultados)
                    return
                else:
                    resultados.append(("!", "NIVEL 3 FALLO — Nombre en cache pero archivo ya no existe",
                        "Cache apunta a \"{}\" pero ese archivo no esta en la carpeta.\n"
                        "Ejecuta 'Indexar tiendas' para actualizar el cache.".format(clave_dest), YELLOW))
            else:
                resultados.append(("X", "NIVEL 3 FALLO — No aparece en el cache col E",
                    "El nombre \"{}\" no esta en el cache ({} nombres indexados).\n"
                    "Causas posibles:\n"
                    "  - Nunca se ejecuto 'Indexar tiendas' o el cache esta desactualizado\n"
                    "  - El nombre en col E del archivo destino es diferente\n"
                    "  - La celda usa formula o formato especial que el indexador no captura".format(
                        nombre_raw, len(indice_cole)), RED))

            # ── CONCLUSION ─────────────────────────────────────────────────
            resultados.append(None)
            resultados.append((">>", "CONCLUSION — Los 3 niveles fallaron",
                "Soluciones:\n"
                "  1. Renombrar el .xlsx para que coincida con \"{}\"\n"
                "  2. Agregar alias en hoja TIENDAS: col A=\"{}\"  col B=nombre del archivo\n"
                "  3. Verificar que \"{}\" aparece en col E del archivo y re-indexar\n"
                "  4. Revisar espacios invisibles o caracteres especiales en el nombre del INGRESO".format(
                    nombre_raw, nombre_raw, nombre_raw), ORANGE))

        except Exception as e:
            resultados = [("X", "Error inesperado en el diagnostico", str(e), RED)]

        finally:
            self.after(0, self._render_diagnostico, frame, resultados)

    def _diag_leer_hoja_tiendas_zip(self, ruta_xlsx, nombre_hoja, fila_ini, norm_fn):
        """Lee la hoja TIENDAS directo del ZIP — sin openpyxl para evitar bloqueos."""
        import zipfile as _zf, re as _re
        mapa = {}
        _PAT_SS = _re.compile(rb'<si>.*?<t[^>]*>(.*?)</t>.*?</si>', _re.DOTALL)
        _PAT_C  = _re.compile(
            rb'<c r="([A-Z]+)(\d+)"[^>]*>'
            rb'(?:<is><t[^>]*>(.*?)</t></is>|<v>(.*?)</v>)',
            _re.DOTALL)
        with _zf.ZipFile(ruta_xlsx, 'r') as z:
            nz = z.namelist()
            strings = []
            for ss in ['xl/sharedStrings.xml', 'xl/SharedStrings.xml']:
                if ss in nz:
                    strings = [m.group(1).decode('utf-8','ignore').strip()
                               for m in _PAT_SS.finditer(z.read(ss))]
                    break
            # Encontrar la hoja correcta por nombre via workbook.xml.rels
            hoja_xml = None
            wb_rels = None
            for cand in ['xl/workbook.xml', 'xl/Workbook.xml']:
                if cand in nz:
                    wb_rels = z.read(cand).decode('utf-8','ignore')
                    break
            if wb_rels:
                # Buscar sheetId de la hoja TIENDAS
                pat_sheet = _re.compile(r'<sheet\s[^>]*name="([^"]+)"[^>]*r:id="([^"]+)"')
                rels_xml  = None
                for cand in ['xl/_rels/workbook.xml.rels', 'xl/_rels/Workbook.xml.rels']:
                    if cand in nz:
                        rels_xml = z.read(cand).decode('utf-8','ignore')
                        break
                if rels_xml:
                    for m in pat_sheet.finditer(wb_rels):
                        if m.group(1).strip().upper() == nombre_hoja.upper():
                            rid = m.group(2)
                            pat_rel = _re.compile(r'Id="' + _re.escape(rid) + r'"[^>]*Target="([^"]+)"')
                            rm = pat_rel.search(rels_xml)
                            if rm:
                                target = rm.group(1).replace("../","xl/").replace("./","xl/worksheets/")
                                if target in nz:
                                    hoja_xml = target
                                    break
            if not hoja_xml:
                # Fallback: primera hoja
                for cand in nz:
                    if 'worksheets/sheet' in cand and cand.endswith('.xml'):
                        hoja_xml = cand; break
            if not hoja_xml:
                return mapa
            raw = z.read(hoja_xml)

        def _val(inline, v_raw):
            if inline:
                return inline.decode('utf-8','ignore').strip()
            if v_raw:
                s = v_raw.decode('utf-8','ignore').strip()
                if s.isdigit() and strings:
                    idx = int(s)
                    return strings[idx] if idx < len(strings) else s
                return s
            return ""

        filas = {}
        for m in _PAT_C.finditer(raw):
            col_l  = m.group(1).decode()
            fila_n = int(m.group(2))
            if fila_n < fila_ini or col_l not in ('A','B'):
                continue
            v = _val(m.group(3), m.group(4))
            if fila_n not in filas:
                filas[fila_n] = {}
            filas[fila_n][col_l] = v

        for datos in filas.values():
            ka = norm_fn(datos.get('A',''))
            kb = norm_fn(datos.get('B',''))
            if ka and kb and ka not in mapa:
                mapa[ka] = kb
        return mapa

    def _diag_verificar_hoja_zip(self, ruta_archivo, encabezados_validos,
                                   fila_encabezado, resultados):
        """Verifica encabezados leyendo el ZIP directo — sin openpyxl."""
        import zipfile as _zf, re as _re
        try:
            _PAT_SS = _re.compile(rb'<si>.*?<t[^>]*>(.*?)</t>.*?</si>', _re.DOTALL)
            _PAT_C  = _re.compile(
                rb'<c r="([A-Z]+)(\d+)"[^>]*>'
                rb'(?:<is><t[^>]*>(.*?)</t></is>|<v>(.*?)</v>)',
                _re.DOTALL)

            def _col_letra_a_num(letra):
                n = 0
                for c in letra:
                    n = n * 26 + (ord(c) - ord('A') + 1)
                return n

            with _zf.ZipFile(ruta_archivo, 'r') as z:
                nz = z.namelist()
                strings = []
                for ss in ['xl/sharedStrings.xml','xl/SharedStrings.xml']:
                    if ss in nz:
                        strings = [m.group(1).decode('utf-8','ignore').strip()
                                   for m in _PAT_SS.finditer(z.read(ss))]
                        break
                hoja_xml = None
                for cand in nz:
                    if 'worksheets/sheet' in cand and cand.endswith('.xml'):
                        hoja_xml = cand; break
                if not hoja_xml:
                    resultados.append(("X", "No se encontro ninguna hoja XML dentro del archivo", "", RED))
                    return
                raw = z.read(hoja_xml)

            def _val(inline, v_raw):
                if inline:
                    return inline.decode('utf-8','ignore').strip()
                if v_raw:
                    s = v_raw.decode('utf-8','ignore').strip()
                    if s.isdigit() and strings:
                        idx = int(s)
                        return strings[idx] if idx < len(strings) else s
                    return s
                return ""

            # Extraer celdas de la fila de encabezado
            enc_encontrados = {}
            for m in _PAT_C.finditer(raw):
                col_l  = m.group(1).decode()
                fila_n = int(m.group(2))
                if fila_n != fila_encabezado:
                    continue
                col_n = _col_letra_a_num(col_l)
                enc_encontrados[col_n] = _val(m.group(3), m.group(4))

            errores = []
            for col, esperado in encabezados_validos.items():
                actual = enc_encontrados.get(col, "")
                if actual.strip().upper() != esperado.upper():
                    errores.append("Col {}: esperado \"{}\" — encontrado \"{}\"".format(
                        col, esperado, actual or "vacio"))

            if not errores:
                resultados.append(("OK", "Hoja valida — encabezados correctos en fila {}".format(fila_encabezado),
                    "El archivo puede recibir paquetes sin problema.", GREEN))
            else:
                resultados.append(("X", "Archivo encontrado pero encabezados incorrectos — NO puede recibir paquetes",
                    "Diferencias en fila {}:\n  ".format(fila_encabezado) + "\n  ".join(errores), RED))

        except Exception as ex:
            resultados.append(("X", "Error al inspeccionar el archivo",
                str(ex), RED))

    def _diag_similares(self, norm, claves, n=4):
        def similitud(a, b):
            if not a or not b:
                return 0.0
            set_a, set_b = set(a), set(b)
            return len(set_a & set_b) / max(len(set_a | set_b), 1)
        scored = sorted(claves, key=lambda k: similitud(norm, k), reverse=True)
        # Solo devolver los que tengan similitud > 0
        return [k for k in scored[:n] if similitud(norm, k) > 0.2]

    def _render_diagnostico(self, frame, resultados):
        for w in frame.winfo_children():
            w.destroy()
        for item in resultados:
            if item is None:
                tk.Frame(frame, bg=BORDER2, height=1).pack(fill="x", padx=12, pady=6)
                continue
            icono, titulo, detalle, color = item
            self._diag_row(frame, icono, titulo, detalle, color)


    # =========================================================================
    # HELPERS COMPARTIDOS — leer/escribir hojas del INGRESO_MASIVO.xlsx
    # =========================================================================
    def _get_ruta_ingreso(self):
        """Retorna la ruta del INGRESO_MASIVO desde config, o None si no existe."""
        import importlib.util
        try:
            spec = importlib.util.spec_from_file_location("config_local", CONFIG_PY)
            cfg  = importlib.util.module_from_spec(spec)
            spec.loader.exec_module(cfg)
            return cfg.ARCHIVO_INGRESO
        except Exception:
            return None

    def _leer_hoja_excel(self, ruta, nombre_hoja, n_cols):
        """
        Abre el Excel y lee la hoja indicada desde fila 2.
        n_cols: cuántas columnas leer por fila (1 para blacklist, 2 para omisiones).
        Retorna lista de tuplas con los valores, sin None.
        """
        from openpyxl import load_workbook
        filas = []
        try:
            wb = load_workbook(ruta, read_only=True, data_only=True)
            if nombre_hoja not in wb.sheetnames:
                wb.close()
                return []
            ws = wb[nombre_hoja]
            for row in ws.iter_rows(min_row=2, max_col=n_cols, values_only=True):
                vals = tuple(str(c).strip() if c is not None else "" for c in row)
                if any(v for v in vals):
                    filas.append(vals)
            wb.close()
        except Exception:
            pass
        return filas

    def _guardar_hoja_excel(self, ruta, nombre_hoja, encabezados, filas):
        """
        Abre el Excel, borra la hoja indicada y la recrea con los nuevos datos.
        encabezados: lista de strings para la fila 1.
        filas: lista de tuplas con los valores a escribir desde fila 2.
        """
        from openpyxl import load_workbook
        try:
            wb = load_workbook(ruta, keep_vba=False)
            if nombre_hoja in wb.sheetnames:
                del wb[nombre_hoja]
            ws = wb.create_sheet(nombre_hoja)
            for ci, enc in enumerate(encabezados, start=1):
                ws.cell(row=1, column=ci).value = enc
            for ri, fila in enumerate(filas, start=2):
                for ci, val in enumerate(fila, start=1):
                    ws.cell(row=ri, column=ci).value = val
            wb.save(ruta)
            wb.close()
            return True, ""
        except Exception as e:
            return False, str(e)

    # =========================================================================
    # LISTA DE OMISIONES  (hoja OMISIONES en INGRESO_MASIVO.xlsx)
    # =========================================================================
    def _abrir_omisiones(self):
        import glob as _g

        ruta = self._get_ruta_ingreso()
        if not ruta or not os.path.isfile(ruta):
            messagebox.showerror("Error",
                "No se encontró INGRESO_MASIVO.\nConfigura la ruta en 'Cambiar rutas'.",
                parent=self)
            return

        # Leer datos actuales desde la hoja Excel
        filas_actuales = self._leer_hoja_excel(ruta, "OMISIONES", 2)

        # Archivos disponibles para autocompletado
        try:
            import importlib.util
            spec = importlib.util.spec_from_file_location("config_local", CONFIG_PY)
            cfg  = importlib.util.module_from_spec(spec)
            spec.loader.exec_module(cfg)
            archivos_disp = sorted([
                os.path.splitext(os.path.basename(r))[0]
                for r in _g.glob(os.path.join(cfg.CARPETA_TIENDAS, "*.xls*"))
                if not os.path.basename(r).startswith("~$")
            ])
        except Exception:
            archivos_disp = []

        dlg = tk.Toplevel(self)
        dlg.title("Lista de Omisiones")
        dlg.geometry("700x520")
        dlg.configure(bg=BG)
        dlg.resizable(True, True)
        dlg.grab_set()
        dlg.transient(self)

        tk.Frame(dlg, bg=ORANGE3, height=3).pack(fill="x")

        hdr = tk.Frame(dlg, bg=BG2)
        hdr.pack(fill="x")
        tk.Label(hdr, text="Lista de Omisiones", bg=BG2, fg=WHITE,
                 font=("Segoe UI",13,"bold")).pack(side="left", padx=20, pady=12)
        tk.Label(hdr,
                 text="Fuerza el ingreso ignorando la búsqueda automática",
                 bg=BG2, fg=WHITE3, font=UI_SM).pack(side="left", padx=4)

        # Indicador de fuente
        tk.Label(hdr, text="● Excel", bg=BG2, fg=GREEN,
                 font=("Segoe UI",7,"bold")).pack(side="right", padx=16)
        tk.Frame(dlg, bg=BORDER, height=1).pack(fill="x")

        tk.Label(dlg,
                 text="  Columna E (como aparece en INGRESO_MASIVO)  →  Nombre del archivo de tienda (sin .xlsx)",
                 bg=BG, fg=WHITE4, font=("Segoe UI",8)).pack(anchor="w", padx=16, pady=(8,4))

        # Encabezados tabla
        enc = tk.Frame(dlg, bg=BG3)
        enc.pack(fill="x", padx=16)
        tk.Label(enc, text="Nombre en Columna E", bg=BG3, fg=WHITE3,
                 font=("Segoe UI",8,"bold"), width=30, anchor="w").pack(
                     side="left", padx=8, pady=5)
        tk.Label(enc, text="Nombre del archivo destino (sin .xlsx)", bg=BG3,
                 fg=WHITE3, font=("Segoe UI",8,"bold"), anchor="w").pack(
                     side="left", padx=4, pady=5)

        # Área scrollable
        outer = tk.Frame(dlg, bg=BG)
        outer.pack(fill="both", expand=True, padx=16, pady=4)
        cvs = tk.Canvas(outer, bg=BG, highlightthickness=0)
        vsb = tk.Scrollbar(outer, orient="vertical", command=cvs.yview)
        cvs.configure(yscrollcommand=vsb.set)
        vsb.pack(side="right", fill="y")
        cvs.pack(side="left", fill="both", expand=True)
        body = tk.Frame(cvs, bg=BG)
        win  = cvs.create_window((0,0), window=body, anchor="nw")
        body.bind("<Configure>",
                  lambda e: cvs.configure(scrollregion=cvs.bbox("all")))
        cvs.bind("<Configure>",
                 lambda e: cvs.itemconfig(win, width=e.width))
        cvs.bind("<MouseWheel>",
                 lambda e: cvs.yview_scroll(int(-1*(e.delta/120)), "units"))

        filas_om = []

        def _agregar_fila(col_e="", col_arch=""):
            fila_f = tk.Frame(body, bg=BG)
            fila_f.pack(fill="x", pady=2)

            var_e    = tk.StringVar(value=col_e)
            var_arch = tk.StringVar(value=col_arch)

            tk.Entry(fila_f, textvariable=var_e, bg=BG4, fg=WHITE,
                     font=UI_SM, relief="flat", bd=0, width=30,
                     highlightbackground=BORDER2, highlightthickness=1,
                     insertbackground=ORANGE).pack(side="left", ipady=5, padx=(0,6))

            ent_arch = tk.Entry(fila_f, textvariable=var_arch, bg=BG4, fg=GREEN,
                                font=UI_SM, relief="flat", bd=0, width=38,
                                highlightbackground=BORDER2, highlightthickness=1,
                                insertbackground=GREEN)
            ent_arch.pack(side="left", ipady=5, padx=(0,4), fill="x", expand=True)

            def _autocomplete(e, ent=ent_arch, var=var_arch):
                txt = var.get().strip().lower()
                if not txt or len(txt) < 2:
                    return
                matches = [a for a in archivos_disp if txt in a.lower()]
                if len(matches) == 1:
                    var.set(matches[0])
                    ent.icursor("end")

            ent_arch.bind("<FocusOut>", _autocomplete)

            idx = len(filas_om)
            def _del(i=idx):
                try:
                    filas_om[i]["frame"].destroy()
                    filas_om[i] = None
                except Exception:
                    pass

            tk.Button(fila_f, text="✕", bg=BG, fg=RED,
                      font=("Segoe UI",9), relief="flat", cursor="hand2",
                      bd=0, padx=6, activebackground=BG3,
                      command=_del).pack(side="left", padx=2)

            filas_om.append({"frame": fila_f, "e": var_e, "arch": var_arch})

        for vals in filas_actuales:
            _agregar_fila(vals[0] if len(vals) > 0 else "",
                          vals[1] if len(vals) > 1 else "")

        tk.Label(dlg,
                 text="Archivos disponibles: {}  |  Guardado en: hoja OMISIONES del INGRESO_MASIVO.xlsx".format(
                     len(archivos_disp)),
                 bg=BG, fg=WHITE4, font=("Segoe UI",7)).pack(anchor="w", padx=16, pady=(0,2))

        # Botones
        bf = tk.Frame(dlg, bg=BG)
        bf.pack(fill="x", padx=16, pady=10)

        tk.Button(bf, text="+ Agregar fila", bg=BG3, fg=ORANGE3,
                  font=UI_SM, relief="flat", cursor="hand2",
                  bd=0, padx=12, pady=7, activebackground=BG4,
                  command=lambda: _agregar_fila()).pack(side="left", padx=(0,8))

        def _guardar():
            nuevas = []
            for r in filas_om:
                if r is None:
                    continue
                k = r["e"].get().strip()
                v = r["arch"].get().strip()
                if k and v:
                    nuevas.append((k, v))

            ok, err = self._guardar_hoja_excel(
                ruta, "OMISIONES",
                ["NOMBRE_TIENDA", "ARCHIVO_DESTINO"],
                nuevas)

            if ok:
                self._log_add("✓ Omisiones guardadas en Excel ({} entradas).".format(
                    len(nuevas)), "ok")
                dlg.destroy()
            else:
                messagebox.showerror("Error",
                    "No se pudo guardar en Excel:\n{}".format(err), parent=dlg)

        tk.Button(bf, text="Cancelar", bg=BG3, fg=WHITE3,
                  font=UI_SM, relief="flat", cursor="hand2",
                  bd=0, padx=14, pady=7, activebackground=BG4,
                  command=dlg.destroy).pack(side="left")

        tk.Button(bf, text="  Guardar omisiones  ",
                  bg=ORANGE3, fg=WHITE, font=UI_B,
                  relief="flat", cursor="hand2", bd=0,
                  padx=18, pady=7, activebackground=ORANGE2,
                  command=_guardar).pack(side="right")

    # =========================================================================
    # LISTA NEGRA  (hoja BLACKLIST en INGRESO_MASIVO.xlsx)
    # =========================================================================
    def _abrir_blacklist(self):
        ruta = self._get_ruta_ingreso()
        if not ruta or not os.path.isfile(ruta):
            messagebox.showerror("Error",
                "No se encontró INGRESO_MASIVO.\nConfigura la ruta en 'Cambiar rutas'.",
                parent=self)
            return

        # Leer datos actuales desde la hoja Excel
        filas_actuales = self._leer_hoja_excel(ruta, "BLACKLIST", 1)

        dlg = tk.Toplevel(self)
        dlg.title("Lista Negra — Tiendas Bloqueadas")
        dlg.geometry("520x500")
        dlg.configure(bg=BG)
        dlg.resizable(True, True)
        dlg.grab_set()
        dlg.transient(self)

        tk.Frame(dlg, bg=RED, height=3).pack(fill="x")

        hdr = tk.Frame(dlg, bg=BG2)
        hdr.pack(fill="x")
        tk.Label(hdr, text="Lista Negra", bg=BG2, fg=WHITE,
                 font=("Segoe UI",13,"bold")).pack(side="left", padx=20, pady=12)
        tk.Label(hdr, text="Tiendas bloqueadas — jamás se ingresa en ellas",
                 bg=BG2, fg=WHITE3, font=UI_SM).pack(side="left", padx=4)

        # Indicador de fuente
        tk.Label(hdr, text="● Excel", bg=BG2, fg=GREEN,
                 font=("Segoe UI",7,"bold")).pack(side="right", padx=16)
        tk.Frame(dlg, bg=BORDER, height=1).pack(fill="x")

        tk.Label(dlg,
                 text="  Escribe el nombre exactamente como aparece en la columna E del INGRESO_MASIVO.",
                 bg=BG, fg=WHITE4, font=("Segoe UI",8)).pack(anchor="w", padx=16, pady=(8,2))
        tk.Label(dlg,
                 text="  Las filas de estas tiendas se marcarán BLOQ y no se procesarán.",
                 bg=BG, fg=RED, font=("Segoe UI",8)).pack(anchor="w", padx=16, pady=(0,4))

        # Encabezado tabla
        enc = tk.Frame(dlg, bg=BG3)
        enc.pack(fill="x", padx=16)
        tk.Label(enc, text="Tienda bloqueada (nombre en Columna E)", bg=BG3, fg=WHITE3,
                 font=("Segoe UI",8,"bold"), anchor="w").pack(
                     side="left", padx=8, pady=5)
        lbl_cnt = tk.Label(enc, text="", bg=BG3, fg=RED,
                           font=("Segoe UI",8,"bold"))
        lbl_cnt.pack(side="right", padx=8)

        # Área scrollable
        outer = tk.Frame(dlg, bg=BG)
        outer.pack(fill="both", expand=True, padx=16, pady=4)
        cvs = tk.Canvas(outer, bg=BG, highlightthickness=0)
        vsb = tk.Scrollbar(outer, orient="vertical", command=cvs.yview)
        cvs.configure(yscrollcommand=vsb.set)
        vsb.pack(side="right", fill="y")
        cvs.pack(side="left", fill="both", expand=True)
        body = tk.Frame(cvs, bg=BG)
        win  = cvs.create_window((0,0), window=body, anchor="nw")
        body.bind("<Configure>",
                  lambda e: cvs.configure(scrollregion=cvs.bbox("all")))
        cvs.bind("<Configure>",
                 lambda e: cvs.itemconfig(win, width=e.width))
        cvs.bind("<MouseWheel>",
                 lambda e: cvs.yview_scroll(int(-1*(e.delta/120)), "units"))

        filas_bl = []

        def _actualizar_cnt():
            n = sum(1 for r in filas_bl if r is not None)
            lbl_cnt.configure(text="{} bloqueada{}".format(n, "s" if n != 1 else ""))

        def _agregar_fila(tienda=""):
            fila_f = tk.Frame(body, bg=BG4,
                              highlightbackground=BORDER2, highlightthickness=1)
            fila_f.pack(fill="x", pady=2)

            # Ícono de bloqueo
            tk.Label(fila_f, text="🚫", bg=BG4, fg=RED,
                     font=("Segoe UI",10)).pack(side="left", padx=(8,4))

            var = tk.StringVar(value=tienda)
            tk.Entry(fila_f, textvariable=var, bg=BG4, fg=RED,
                     font=UI_SM, relief="flat", bd=0,
                     highlightthickness=0,
                     insertbackground=RED).pack(
                         side="left", ipady=6, fill="x", expand=True)

            idx = len(filas_bl)
            def _del(i=idx):
                try:
                    filas_bl[i]["frame"].destroy()
                    filas_bl[i] = None
                    _actualizar_cnt()
                except Exception:
                    pass

            tk.Button(fila_f, text="✕", bg=BG4, fg=WHITE4,
                      font=("Segoe UI",9), relief="flat", cursor="hand2",
                      bd=0, padx=8, pady=4, activebackground=BG3,
                      command=_del).pack(side="right", padx=4)

            filas_bl.append({"frame": fila_f, "v": var})
            _actualizar_cnt()

        for vals in filas_actuales:
            _agregar_fila(vals[0] if vals else "")

        _actualizar_cnt()

        tk.Label(dlg,
                 text="Guardado en: hoja BLACKLIST del INGRESO_MASIVO.xlsx  |  Tolerante a mayúsculas y tildes",
                 bg=BG, fg=WHITE4, font=("Segoe UI",7)).pack(anchor="w", padx=16, pady=(0,2))

        # Botones
        bf = tk.Frame(dlg, bg=BG)
        bf.pack(fill="x", padx=16, pady=10)

        tk.Button(bf, text="+ Agregar tienda", bg=BG3, fg=RED,
                  font=UI_SM, relief="flat", cursor="hand2",
                  bd=0, padx=12, pady=7, activebackground=BG4,
                  command=lambda: _agregar_fila()).pack(side="left", padx=(0,8))

        def _guardar():
            nuevas = []
            for r in filas_bl:
                if r is None:
                    continue
                v = r["v"].get().strip()
                if v:
                    nuevas.append((v,))

            ok, err = self._guardar_hoja_excel(
                ruta, "BLACKLIST",
                ["TIENDA_BLOQUEADA"],
                nuevas)

            if ok:
                self._log_add("✓ Lista negra guardada en Excel ({} tiendas bloqueadas).".format(
                    len(nuevas)), "ok")
                dlg.destroy()
            else:
                messagebox.showerror("Error",
                    "No se pudo guardar en Excel:\n{}".format(err), parent=dlg)

        tk.Button(bf, text="Cancelar", bg=BG3, fg=WHITE3,
                  font=UI_SM, relief="flat", cursor="hand2",
                  bd=0, padx=14, pady=7, activebackground=BG4,
                  command=dlg.destroy).pack(side="left")

        tk.Button(bf, text="  Guardar lista negra  ",
                  bg=RED, fg=WHITE, font=UI_B,
                  relief="flat", cursor="hand2", bd=0,
                  padx=18, pady=7, activebackground="#CC0020",
                  command=_guardar).pack(side="right")

    # =========================================================================
    # SISTEMA DE ACTUALIZACIONES
    # =========================================================================
    def _check_update_bg(self):
        """Corre en background — consulta GitHub y muestra banner si hay update."""
        try:
            hay, remota = _hay_actualizacion()
            if hay and remota:
                self._version_remota = remota
                self.after(0, self._mostrar_banner_update, remota)
        except Exception:
            pass

    def _mostrar_banner_update(self, remota):
        """Muestra el banner verde en el sidebar con la version disponible."""
        v_nueva = remota.get("version", "?")
        v_local = _leer_version_local().get("version", "?")
        self._update_lbl.configure(
            text="🔄 Nueva versión disponible\n"
                 "  {} → {}\n"
                 "  Clic para actualizar".format(v_local, v_nueva))
        self._update_banner.pack(fill="x", padx=10, pady=(0, 6),
                                  before=self._lbl_ver.master)

    def _mostrar_dialogo_update(self):
        """Muestra el dialogo de confirmacion de actualizacion."""
        if not self._version_remota:
            return
        remota  = self._version_remota
        v_nueva = remota.get("version", "?")
        v_local = _leer_version_local().get("version", "?")
        notas   = remota.get("notas", "Sin notas.")
        archivos = remota.get("archivos", [])

        dlg = tk.Toplevel(self)
        dlg.title("Actualización disponible")
        dlg.geometry("420x340")
        dlg.configure(bg=BG2)
        dlg.resizable(False, False)
        dlg.grab_set()
        dlg.transient(self)

        tk.Frame(dlg, bg=GREEN, height=3).pack(fill="x")
        tk.Label(dlg, text="🔄 Actualización disponible",
                 bg=BG2, fg=WHITE, font=("Segoe UI", 12, "bold")).pack(pady=(16, 4))
        tk.Label(dlg, text="v{}  →  v{}".format(v_local, v_nueva),
                 bg=BG2, fg=GREEN, font=("Segoe UI", 11, "bold")).pack()

        tk.Frame(dlg, bg=BORDER, height=1).pack(fill="x", padx=20, pady=10)

        tk.Label(dlg, text="Novedades:", bg=BG2, fg=WHITE3,
                 font=("Segoe UI", 8, "bold")).pack(anchor="w", padx=20)
        tk.Label(dlg, text=notas, bg=BG2, fg=WHITE3,
                 font=("Segoe UI", 8), wraplength=360,
                 justify="left").pack(anchor="w", padx=20, pady=(2, 8))

        tk.Label(dlg, text="Archivos a actualizar: {}".format(
            ", ".join(archivos)),
            bg=BG2, fg=WHITE4, font=("Segoe UI", 7),
            wraplength=360, justify="left").pack(anchor="w", padx=20)

        tk.Frame(dlg, bg=BORDER, height=1).pack(fill="x", padx=20, pady=10)

        # Barra progreso
        prog_c = tk.Canvas(dlg, bg=BG3, height=6, highlightthickness=0)
        prog_c.pack(fill="x", padx=20, pady=(0, 6))
        prog_fill = prog_c.create_rectangle(0, 0, 0, 6, fill=GREEN, outline="")

        def _set_prog(pct):
            w = prog_c.winfo_width()
            prog_c.coords(prog_fill, 0, 0, int(w * pct), 6)
            dlg.update_idletasks()

        log_lbl = tk.Label(dlg, text="", bg=BG2, fg=WHITE3,
                            font=("Segoe UI", 8))
        log_lbl.pack(pady=2)

        bf = tk.Frame(dlg, bg=BG2)
        bf.pack(fill="x", padx=20, pady=(4, 14))
        btn_cancel = tk.Button(bf, text="Cancelar", bg=BG3, fg=WHITE3,
                               font=("Segoe UI", 9), relief="flat",
                               cursor="hand2", bd=0, padx=14, pady=7,
                               activebackground=BG4,
                               command=dlg.destroy)
        btn_cancel.pack(side="left")
        btn_update = tk.Button(bf, text="  Actualizar ahora  ",
                               bg=GREEN, fg=BG,
                               font=("Segoe UI", 10, "bold"), relief="flat",
                               cursor="hand2", bd=0, padx=18, pady=7,
                               activebackground="#00C080")
        btn_update.pack(side="right")

        def _hacer_update():
            btn_update.configure(state="disabled")
            btn_cancel.configure(state="disabled")

            def _worker():
                ok, msg = _aplicar_actualizacion(
                    remota,
                    callback_progreso=lambda p: dlg.after(0, _set_prog, p),
                    callback_log=lambda m: dlg.after(0, log_lbl.configure, {"text": m[-60:]}))

                if ok:
                    dlg.after(0, lambda: (
                        messagebox.showinfo(
                            "Actualización completa",
                            "✓ Versión {} instalada.\n\n"
                            "El programa se reiniciará.".format(v_nueva),
                            parent=dlg),
                        dlg.destroy(),
                        self._reiniciar()))
                else:
                    dlg.after(0, lambda: (
                        messagebox.showerror(
                            "Error", "Algunos archivos fallaron:\n{}".format(msg),
                            parent=dlg),
                        btn_cancel.configure(state="normal")))

            threading.Thread(target=_worker, daemon=True).start()

        btn_update.configure(command=_hacer_update)

    def _reiniciar(self):
        """Reinicia el programa para aplicar los nuevos archivos."""
        try:
            base = os.path.dirname(os.path.abspath(__file__))
            # Buscar el launcher o el propio script
            launcher = os.path.join(base, "launcher.exe")
            if os.path.isfile(launcher):
                subprocess.Popen([launcher], cwd=base)
            else:
                subprocess.Popen([sys.executable, __file__], cwd=base)
        except Exception:
            pass
        self.destroy()


if __name__ == "__main__":
    try:
        App().mainloop()
    except Exception as e:
        _mostrar_error(e)
