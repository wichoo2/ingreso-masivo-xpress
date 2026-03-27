import os
import re
import sys
import time
import glob
import copy
import traceback
from collections import defaultdict
from openpyxl import load_workbook

# Forzar stdout en UTF-8 para evitar UnicodeEncodeError en Windows
if hasattr(sys.stdout, 'reconfigure'):
    try:
        sys.stdout.reconfigure(encoding='utf-8', errors='replace')
    except Exception:
        pass

import logica_local as logica
from config_local import (
    CARPETA_TIENDAS, ARCHIVO_INGRESO,
    HOJA_ORIGEN, HOJA_TIENDAS,
    FILA_INICIO, FILA_DATOS_DEST, FILA_TIENDAS_INI,
    COL_RESULTADO, COL_DATOS_INI, COL_TIENDA, COL_ID,
    COL_TIPOSERV, COL_MUNICIPIO, COL_COMENTARIO
)

SEP            = "=" * 60
MAX_LIBROS     = 25
MAX_REINTENTOS = 5
ESPERA_REINT   = 3

import zipfile as _zipfile_main
import deshacer as _deshacer

# =============================================================================
# LISTA DE OMISIONES  (hoja "OMISIONES" en INGRESO_MASIVO.xlsx)
# Columna A: nombre tal como aparece en col E del INGRESO_MASIVO
# Columna B: nombre exacto del archivo destino (sin extension)
# Fila 1:    encabezados (se salta automaticamente)
# Bypasea toda la logica de busqueda — ingresa directo sin validar.
# Al ser parte del Excel compartido en OneDrive todos editan el mismo archivo.
# =============================================================================
HOJA_OMISIONES = "OMISIONES"

def cargar_omisiones(wb):
    """
    Lee la hoja OMISIONES del workbook ya abierto.
    Retorna dict: {nombre_normalizado_col_a: clave_archivo_normalizada}
    """
    if HOJA_OMISIONES not in wb.sheetnames:
        return {}
    ws = wb[HOJA_OMISIONES]
    resultado = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        k = row[0] if len(row) > 0 else None
        v = row[1] if len(row) > 1 else None
        if k is None or v is None:
            continue
        ks = str(k).strip()
        vs = str(v).strip()
        if not ks or not vs:
            continue
        nk = logica.normalizar(ks)
        nv = logica.normalizar(vs)
        if nk and nv:
            resultado[nk] = nv
    return resultado

def _crear_hoja_omisiones(wb):
    """
    NO crea la hoja automaticamente.
    Si el usuario quiere omisiones, la agrega manualmente desde la aplicacion.
    Esta funcion se conserva por compatibilidad pero no hace nada.
    """
    pass

# =============================================================================
# LISTA NEGRA  (hoja "BLACKLIST" en INGRESO_MASIVO.xlsx)
# Columna A: nombre de la tienda a bloquear (tolerante a mayusculas/tildes)
# Fila 1:    encabezado (se salta automaticamente)
# Las filas bloqueadas se marcan "BLOQ" en col B (distinguible de FALTA/DUP).
# Al ser parte del Excel compartido en OneDrive todos editan el mismo archivo.
# =============================================================================
HOJA_BLACKLIST = "BLACKLIST"

def cargar_blacklist(wb):
    """
    Lee la hoja BLACKLIST del workbook ya abierto.
    Retorna set de nombres normalizados bloqueados.
    """
    if HOJA_BLACKLIST not in wb.sheetnames:
        return set()
    ws = wb[HOJA_BLACKLIST]
    resultado = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        v = row[0] if len(row) > 0 else None
        if v is None:
            continue
        vs = str(v).strip()
        if not vs:
            continue
        n = logica.normalizar(vs)
        if n:
            resultado.add(n)
    return resultado

def _crear_hoja_blacklist(wb):
    """
    NO crea la hoja automaticamente.
    Si el usuario quiere blacklist, la agrega manualmente desde la aplicacion.
    Esta funcion se conserva por compatibilidad pero no hace nada.
    """
    pass

# =============================================================================
# LOG PERSISTENTE
# =============================================================================
_LOG_PATH = None

def _init_log():
    global _LOG_PATH
    try:
        base = os.path.dirname(os.path.abspath(__file__))
        _LOG_PATH = os.path.join(base, "ultimo_proceso.log")
        with open(_LOG_PATH, "w", encoding="utf-8") as f:
            f.write("=== INGRESO MASIVO - LOG {} ===\n".format(
                time.strftime("%d/%m/%Y %H:%M:%S")))
    except Exception:
        _LOG_PATH = None

def _log(msg):
    try:
        print(msg)
    except UnicodeEncodeError:
        # Windows cp1252 no soporta algunos caracteres unicode — imprimir seguro
        print(msg.encode('ascii', errors='replace').decode('ascii'))
    if _LOG_PATH:
        try:
            with open(_LOG_PATH, "a", encoding="utf-8") as f:
                f.write(msg + "\n")
        except Exception:
            pass

# =============================================================================
# ONEDRIVE SYNC
# =============================================================================
def _forzar_sync_onedrive(ruta):
    try:
        _now = time.time()
        os.utime(ruta, (_now, _now))
    except Exception:
        pass
    try:
        with open(ruta, "ab") as f:
            pass
    except Exception:
        pass

# =============================================================================
# DETECTAR ARCHIVOS BLOQUEADOS
# =============================================================================

def _verificar_sincronizado(ruta):
    """
    Verifica que el archivo no esté abierto en Excel ni parcialmente
    descargado de OneDrive. Retorna (ok, mensaje).
    """
    try:
        directorio = os.path.dirname(ruta)
        nombre     = os.path.basename(ruta)
        lock_path  = os.path.join(directorio, "~$" + nombre)
        if os.path.isfile(lock_path):
            return False, "abierto en Excel (~$ lock)"
        size = os.path.getsize(ruta)
        if size < 5000:
            return False, "no sincronizado ({} bytes)".format(size)
        try:
            with open(ruta, "rb") as _f:
                _f.read(512)
        except PermissionError:
            return False, "siendo sincronizado por OneDrive"
        return True, ""
    except Exception as e:
        return False, str(e)

def _archivo_bloqueado(ruta):
    directorio = os.path.dirname(ruta)
    nombre     = os.path.basename(ruta)
    lock       = os.path.join(directorio, "~$" + nombre)
    return os.path.isfile(lock)

def _verificar_bloqueados(indice):
    return [os.path.basename(ruta)
            for ruta in indice.values()
            if _archivo_bloqueado(ruta)]

# =============================================================================
# PRESERVAR ZIP (solo tema) — trabaja con BytesIO, sin archivos temporales
# =============================================================================
def _preservar_tema_en_buf(ruta_orig, buf):
    """
    Preserva xl/theme/theme1.xml del archivo original dentro del buffer BytesIO.
    Modifica el buffer en memoria sin tocar el disco.
    """
    import io as _io
    TEMA = "xl/theme/theme1.xml"
    try:
        if not os.path.isfile(ruta_orig):
            return
        with _zipfile_main.ZipFile(ruta_orig, 'r') as z:
            if TEMA not in z.namelist():
                return
            tema_orig = z.read(TEMA)

        buf.seek(0)
        with _zipfile_main.ZipFile(buf, 'r') as z:
            if TEMA not in z.namelist():
                return
            if z.read(TEMA) == tema_orig:
                return  # identicos

        # Reconstruir con tema original
        buf.seek(0)
        datos_zip = {}
        infos_zip = {}
        with _zipfile_main.ZipFile(buf, 'r') as z_in:
            for item in z_in.infolist():
                datos_zip[item.filename] = z_in.read(item.filename)
                infos_zip[item.filename] = item
        datos_zip[TEMA] = tema_orig

        buf_nuevo = _io.BytesIO()
        with _zipfile_main.ZipFile(buf_nuevo, 'w',
                                    compression=_zipfile_main.ZIP_DEFLATED) as z_out:
            for fn, data in datos_zip.items():
                z_out.writestr(infos_zip[fn], data)

        # Reemplazar contenido del buffer original
        contenido = buf_nuevo.getvalue()
        buf.seek(0)
        buf.truncate()
        buf.write(contenido)
        buf.seek(0)
    except Exception:
        pass

# =============================================================================
# CACHE DE LIBROS
# =============================================================================
class CacheLibros:
    def __init__(self):
        self._libros  = {}
        self._orden   = []
        self._resumen = {}
        # Sistema de deshacer — backup antes de tocar cada archivo
        self._sesion_ts, self._carpeta_bak = None, None
        self._archivos_originales = []  # rutas originales modificadas

    def iniciar_sesion_backup(self):
        """Llama al inicio del proceso para preparar la carpeta de backup."""
        self._carpeta_bak, self._sesion_ts = _deshacer.iniciar_sesion()
        self._archivos_originales = []
        return self._sesion_ts

    def tiene(self, clave):      return clave in self._libros
    def obtener(self, clave):    return self._libros.get(clave, {})\


    def count(self):             return len(self._libros)
    def get_resumen(self):       return self._resumen

    def agregar(self, clave, datos):
        self._libros[clave] = datos
        self._orden.append(clave)
        if clave not in self._resumen:
            self._resumen[clave] = {
                "nombre": os.path.basename(datos["path"]),
                "listo": 0, "dup": 0}
        # Backup antes de cualquier modificación
        if self._carpeta_bak:
            ruta_orig = datos["path"]
            _deshacer.guardar_backup(ruta_orig, self._carpeta_bak)
            if ruta_orig not in self._archivos_originales:
                self._archivos_originales.append(ruta_orig)

    def sumar(self, clave, tipo):
        if clave in self._resumen:
            self._resumen[clave][tipo] = self._resumen[clave].get(tipo, 0) + 1

    def cerrar_todos(self):
        import io as _io
        for clave, datos in list(self._libros.items()):
            nombre    = os.path.basename(datos["path"])
            ruta_orig = datos["path"]

            guardado = False
            for intento in range(1, MAX_REINTENTOS + 1):
                try:
                    # Guardar en buffer de memoria — sin crear ningun archivo
                    buf = _io.BytesIO()
                    datos["wb"].save(buf)
                    datos["wb"].close()
                    contenido = buf.getvalue()

                    if contenido:
                        # Preservar tema de colores del original
                        _preservar_tema_en_buf(ruta_orig, buf)
                        contenido = buf.getvalue()

                        # FIX 6: Verificar que el buffer es un ZIP válido
                        # antes de sobreescribir el archivo original
                        import io as _io2
                        buf_check = _io2.BytesIO(contenido)
                        if not _zipfile_main.is_zipfile(buf_check):
                            _log("[ERROR] Buffer de '{}' no es un ZIP válido — "
                                 "no se sobreescribe el original.".format(nombre))
                            break

                        # Verificar sync antes de escribir
                        _ok_sync, _msg_sync = _verificar_sincronizado(ruta_orig)
                        if not _ok_sync:
                            _log("[SYNC] '{}': {} — reintentando en {}s...".format(
                                nombre, _msg_sync, ESPERA_REINT))
                            raise PermissionError(_msg_sync)
                        # Escribir directamente sobre el archivo original
                        with open(ruta_orig, 'wb') as f:
                            f.write(contenido)
                        _forzar_sync_onedrive(ruta_orig)
                        _log("[OK] Guardado: {}".format(nombre))
                        guardado = True
                    else:
                        _log("[ERROR] Buffer vacio: {}".format(nombre))
                    break
                except PermissionError:
                    if intento < MAX_REINTENTOS:
                        _log("  [Reintento {}/{}] '{}' bloqueado, esperando {}s...".format(
                            intento, MAX_REINTENTOS, nombre, ESPERA_REINT))
                        time.sleep(ESPERA_REINT)
                    else:
                        _log("[ERROR] No se pudo guardar '{}' despues de {} intentos".format(
                            nombre, MAX_REINTENTOS))
                except Exception as e:
                    _log("[ERROR] Al guardar {}: {}".format(nombre, e))
                    break

            if not guardado:
                _log("[ERROR CRITICO] '{}' NO fue guardado.".format(nombre))

        self._libros.clear()
        self._orden.clear()

# =============================================================================
# INDEXAR CARPETA
# =============================================================================
def indexar_carpeta(carpeta):
    """
    Construye dos indices a partir de los archivos .xlsx de la carpeta:
      indice          -> {normalizar(sin_ext):          ruta}
      indice_sin_tipo -> {normalizar_sin_tipo(sin_ext): ruta}
    El segundo permite el Nivel 4: matching tolerante a sufijos de tipo
    de cuenta como (SCGE), (PLS), LOCKERS, METROGALERIAS, etc.
    """
    indice          = {}
    indice_sin_tipo = {}
    for ruta in glob.glob(os.path.join(carpeta, "*.xls*")):
        nombre = os.path.basename(ruta)
        if nombre.startswith("~$"):
            continue
        dot     = nombre.rfind(".")
        sin_ext = nombre[:dot] if dot > 0 else nombre
        norm    = logica.normalizar(sin_ext)
        norm_st = logica.normalizar_sin_tipo(sin_ext)
        if norm and norm not in indice:
            indice[norm] = ruta
        # Solo agregar al indice_sin_tipo si la clave es diferente
        # (evitar entradas redundantes cuando el nombre no tiene sufijo)
        if norm_st and norm_st not in indice_sin_tipo:
            indice_sin_tipo[norm_st] = ruta
    return indice, indice_sin_tipo

# =============================================================================
# CARGAR LIBRO DESTINO
# =============================================================================
def cargar_libro(ruta, clave):
    # FIX 5: Detectar placeholder de OneDrive ANTES de intentar abrir
    # Un archivo no descargado pesa < 5KB y load_workbook falla con error de ZIP
    try:
        size = os.path.getsize(ruta)
        if size < 5000:
            _log("[SYNC] '{}' no descargado de OneDrive ({} bytes). "
                 "Haz clic derecho → 'Mantener siempre en este dispositivo'.".format(
                     os.path.basename(ruta), size))
            return None
    except Exception:
        pass

    try:
        wb = load_workbook(ruta, keep_vba=False, data_only=False)
    except Exception as e:
        _log("[ERROR] No se pudo abrir {}: {}".format(os.path.basename(ruta), e))
        return None

    ws_valida = None
    for nombre_hoja in wb.sheetnames:
        ws = wb[nombre_hoja]
        if ws.sheet_state != "visible":
            continue
        # FIX 4: Se eliminó el check ws.protection.sheet
        # openpyxl puede escribir en hojas con protección sin contraseña.
        # El check descartaba hojas válidas que solo protegen encabezados.
        if logica.hoja_valida(ws):
            ws_valida = ws
            break

    if ws_valida is None:
        _log("[WARN] '{}' — ninguna hoja visible con encabezados válidos.".format(
            os.path.basename(ruta)))
        wb.close()
        return None

    cols_esp       = logica.detectar_cols_especiales(ws_valida)
    ids            = logica.cargar_ids_destino(ws_valida)
    fila_enc_real  = logica.get_fila_encabezado(ws_valida)
    fila_datos_ini = fila_enc_real + 1

    nombres_cole = set()
    for row in ws_valida.iter_rows(min_row=fila_datos_ini, max_row=ws_valida.max_row,
                                    min_col=5, max_col=5, values_only=True):
        raw = row[0]
        if raw is None:
            continue
        v = str(raw).strip()
        if v.startswith("#") or v == "":
            continue
        n = logica.normalizar(v)
        if n and len(n) >= 3 and not n.isdigit():
            nombres_cole.add(n)

    ult_datos  = logica.ultima_fila_con_datos(ws_valida)
    fila_libre = logica.primera_fila_libre(ws_valida, ult_datos + 1)

    return {
        "wb":           wb,
        "ws":           ws_valida,
        "path":         ruta,
        "tipo_q":       cols_esp["tipo_q"],
        "col_q":        cols_esp["col_q"],
        "col_paq":      cols_esp["col_paq"],
        "col_oid":      cols_esp["col_oid"],
        "ids":          ids,
        "nombres_cole": nombres_cole,
        "clave":        clave,
        "fila_libre":   fila_libre,
        "fila_enc":     fila_enc_real,
    }

# =============================================================================
# PRE-CLASIFICAR FILAS POR TIENDA (evita busquedas repetidas)
# =============================================================================
def _clasificar_filas(filas_data, indice, indice_sin_tipo, mapa_tiendas,
                      indice_cole, ws_origen, ultima_fila, omisiones, blacklist):
    """
    Paso 1: clasifica todas las filas en memoria agrupandolas por tienda.
    Blacklist : bloquea antes que todo — marca BLOQ.
    Omisiones : bypass directo al archivo mapeado manualmente.
    Nivel 1   : normalizar(col E) == normalizar(nombre_archivo)
    Nivel 2   : alias en hoja TIENDAS
    Nivel 3   : cache columna E de los archivos destino
    Nivel 4   : normalizar_sin_tipo() — tolerante a sufijos (SCGE)(PLS) LOCKERS etc.
    """
    _norm_cache    = {}
    _normst_cache  = {}
    # omisiones y blacklist se reciben como parametros (ya leidos del workbook)

    grupos = defaultdict(list)
    faltas = []
    bloqueadas = []

    for fila_real in range(FILA_INICIO, ultima_fila + 1):
        fila_vals = filas_data.get(fila_real, ())

        def _v(col):
            idx = col - 1
            v = fila_vals[idx] if idx < len(fila_vals) else None
            return str(v).strip() if v is not None else ""

        nombre_tienda = _v(COL_TIENDA)
        id_nuevo      = _v(COL_ID)
        arr_datos     = [fila_vals[COL_DATOS_INI - 1 + j]
                         if (COL_DATOS_INI - 1 + j) < len(fila_vals) else None
                         for j in range(7)]

        # Fila completamente vacia
        if (not nombre_tienda and not id_nuevo and
                all(v is None or str(v).strip() == "" for v in arr_datos)):
            continue

        # Sin tienda
        if not nombre_tienda:
            ws_origen.cell(row=fila_real, column=COL_RESULTADO).value = "FALTA"
            faltas.append((fila_real, list(fila_vals), "sin_tienda"))
            continue

        # Normalizar con cache
        if nombre_tienda not in _norm_cache:
            _norm_cache[nombre_tienda]   = logica.normalizar(nombre_tienda)
            _normst_cache[nombre_tienda] = logica.normalizar_sin_tipo(nombre_tienda)
        norm_tienda    = _norm_cache[nombre_tienda]
        norm_tienda_st = _normst_cache[nombre_tienda]

        # [-1] BLACKLIST — maxima prioridad absoluta, bloquea antes que todo
        # Compara tanto la forma exacta como la forma sin sufijo
        if norm_tienda in blacklist or norm_tienda_st in blacklist:
            ws_origen.cell(row=fila_real, column=COL_RESULTADO).value = "BLOQ"
            bloqueadas.append((fila_real, list(fila_vals), nombre_tienda))
            _log("  BLOQ   fila {}: '{}' (blacklist)".format(fila_real, nombre_tienda))
            continue

        clave_archivo = None
        motivos_falta = []  # FIX 7: log detallado por nivel

        # [0] Lista de omisiones — bypass directo (maxima prioridad)
        if norm_tienda in omisiones:
            clave_omision = omisiones[norm_tienda]
            if clave_omision in indice:
                clave_archivo = clave_omision
                _log("  [OMISION] fila {}: '{}' -> '{}'".format(
                    fila_real, nombre_tienda, clave_omision))
            else:
                motivos_falta.append("Omision '{}' no encontrada en índice".format(clave_omision))

        # [1] Exacto por nombre archivo
        if not clave_archivo and norm_tienda in indice:
            clave_archivo = norm_tienda
        elif not clave_archivo:
            motivos_falta.append("Niv1: nombre exacto '{}' no en índice".format(norm_tienda))

        # [2] Hoja TIENDAS
        if not clave_archivo and norm_tienda in mapa_tiendas:
            nd = mapa_tiendas[norm_tienda]
            if nd in indice:
                clave_archivo = nd
            else:
                motivos_falta.append("Niv2: alias '{}' no en índice".format(nd))
        elif not clave_archivo:
            motivos_falta.append("Niv2: sin alias en hoja TIENDAS")

        # [3] Cache col E
        if not clave_archivo and norm_tienda in indice_cole:
            clave_archivo = indice_cole[norm_tienda]
        elif not clave_archivo:
            motivos_falta.append("Niv3: '{}' no en caché col E".format(norm_tienda))

        # [4] Sufijo inteligente — tolerante a (SCGE)(PLS) LOCKERS METROGALERIAS
        if not clave_archivo and norm_tienda_st in indice_sin_tipo:
            ruta_st = indice_sin_tipo[norm_tienda_st]
            dot     = os.path.basename(ruta_st).rfind(".")
            sin_ext = os.path.basename(ruta_st)[:dot] if dot > 0 else os.path.basename(ruta_st)
            clave_st = logica.normalizar(sin_ext)
            if clave_st in indice:
                clave_archivo = clave_st
                _log("  [NIV4] fila {}: '{}' -> '{}'".format(
                    fila_real, nombre_tienda, sin_ext))
        elif not clave_archivo:
            motivos_falta.append("Niv4: sin_tipo '{}' no coincide".format(norm_tienda_st))

        # FIX 8 — [5] Palabras clave: buscar archivos con 2+ palabras en común
        # Cubre nombres con caracteres especiales, abreviaciones y variaciones menores
        if not clave_archivo:
            palabras = [p for p in re.split(r'[^a-z0-9]+', norm_tienda) if len(p) >= 3]
            mejor_clave = None
            mejor_score = 0
            for clave_cand in indice:
                score = sum(1 for p in palabras if p in clave_cand)
                if score >= 2 and score > mejor_score:
                    mejor_score = score
                    mejor_clave = clave_cand
            if mejor_clave:
                clave_archivo = mejor_clave
                _log("  [NIV5] fila {}: '{}' -> '{}' ({} palabras)".format(
                    fila_real, nombre_tienda,
                    os.path.basename(indice[mejor_clave]), mejor_score))
            else:
                motivos_falta.append("Niv5: sin coincidencia por palabras clave")

        if not clave_archivo:
            ws_origen.cell(row=fila_real, column=COL_RESULTADO).value = "FALTA"
            faltas.append((fila_real, list(fila_vals), nombre_tienda))

            # FIX 7: Log detallado + FIX 8: Sugerencias de archivos similares
            _log("  FALTA  fila {}: '{}'".format(fila_real, nombre_tienda))
            for m in motivos_falta:
                _log("         {}".format(m))

            # Mostrar hasta 3 sugerencias por similitud
            palabras_s = [p for p in re.split(r'[^a-z0-9]+', norm_tienda) if len(p) >= 3]
            if palabras_s:
                scored = []
                for clave_c, ruta_c in indice.items():
                    sc = sum(1 for p in palabras_s if p in clave_c)
                    if sc >= 1:
                        scored.append((sc, os.path.basename(ruta_c)))
                scored.sort(key=lambda x: -x[0])
                if scored:
                    sugs = [n for _, n in scored[:3]]
                    _log("         ¿Quisiste decir? {}".format(" | ".join(sugs)))
            continue

        grupos[clave_archivo].append((fila_real, fila_vals))

    return grupos, faltas, bloqueadas

# =============================================================================
# PROCESAR UN GRUPO DE FILAS PARA UNA TIENDA
# =============================================================================
def _procesar_grupo(clave_archivo, filas_grupo, cache, indice,
                    indice_cole, ws_origen):
    """
    Procesa todas las filas de UNA tienda de una sola vez.
    El libro se abre una sola vez y se insertan todos sus paquetes juntos.
    """
    c_listo = c_dup = 0

    # Abrir libro si no esta en cache
    if not cache.tiene(clave_archivo):
        ruta        = indice[clave_archivo]
        nombre_arch = os.path.basename(ruta)
        _log("  Abriendo {}...".format(nombre_arch))
        datos = cargar_libro(ruta, clave_archivo)
        if datos is None:
            # Marcar todas las filas del grupo como FALTA
            for fila_real, fila_vals in filas_grupo:
                ws_origen.cell(row=fila_real, column=COL_RESULTADO).value = "FALTA"
            _log("  FALTA  hoja invalida: '{}'".format(nombre_arch))
            return 0, len(filas_grupo), [(fr, list(fv), nombre_arch)
                                          for fr, fv in filas_grupo]
        cache.agregar(clave_archivo, datos)
        for ne in datos["nombres_cole"]:
            if ne not in indice_cole:
                indice_cole[ne] = clave_archivo

    datos_lib = cache.obtener(clave_archivo)
    ws_dest   = datos_lib["ws"]
    tipo_q    = datos_lib["tipo_q"]
    col_q     = datos_lib["col_q"]
    col_paq   = datos_lib["col_paq"]
    col_oid   = datos_lib["col_oid"]
    ids_dest  = datos_lib["ids"]
    fila_libre = datos_lib["fila_libre"]

    filas_falta_grupo = []

    for fila_real, fila_vals in filas_grupo:
        def _v(col):
            idx = col - 1
            v = fila_vals[idx] if idx < len(fila_vals) else None
            return str(v).strip() if v is not None else ""

        id_nuevo = logica.normalizar_id(fila_vals[COL_ID - 1] if COL_ID - 1 < len(fila_vals) else None)
        valor_k  = _v(COL_TIPOSERV)
        valor_m  = _v(COL_MUNICIPIO)
        valor_n  = _v(COL_COMENTARIO)
        arr_datos = [fila_vals[COL_DATOS_INI - 1 + j]
                     if (COL_DATOS_INI - 1 + j) < len(fila_vals) else None
                     for j in range(7)]

        num_paquetes = max(logica.extraer_num_paquetes(valor_n), 1)
        estado, a_insertar = logica.evaluar_duplicado(
            id_nuevo, valor_n, ids_dest, num_paquetes)

        if estado.startswith("DUP"):
            ws_origen.cell(row=fila_real, column=COL_RESULTADO).value = "DUP"
            cache.sumar(clave_archivo, "dup")
            c_dup += 1
            motivo_dup = {
                "DUP_EXISTE": "ya existe en tienda",
                "DUP_PAQ":    "PAQ pero ID ya existe (no fragmentar)",
                "DUP_SESION": "apareció dos veces en este proceso",
                "DUP_VACIO":  "ID vacío",
            }.get(estado, estado)
            _log("  DUP    fila {}: ID {} ({})".format(fila_real, id_nuevo, motivo_dup))
            continue

        ya_existentes = ids_dest.get(id_nuevo, 0)

        for k in range(a_insertar):
            logica.insertar_paquete(
                ws_dest, fila_libre, arr_datos,
                precio_cero=(ya_existentes > 0 or k > 0),
                valor_k=valor_k, valor_m=valor_m, valor_n=valor_n,
                es_primero=(k == 0),
                tipo_q=tipo_q, col_q=col_q,
                col_paq=col_paq, col_oid=col_oid)
            fila_libre += 1
            fila_libre = logica.primera_fila_libre_rapida(ws_dest, fila_libre)

        if id_nuevo:
            ids_dest[id_nuevo] = ya_existentes + a_insertar

        ws_origen.cell(row=fila_real, column=COL_RESULTADO).value = "LISTO"
        cache.sumar(clave_archivo, "listo")
        c_listo += 1
        _log("  LISTO  fila {}".format(fila_real))

    # Actualizar puntero fila libre en cache
    datos_lib["fila_libre"] = fila_libre

    return c_listo, c_dup, filas_falta_grupo

# =============================================================================
# MAIN
# =============================================================================
def main():
    _init_log()
    t_inicio = time.time()

    _log("\n" + SEP)
    _log("  MOVER PAQUETES - MODO LOCAL")
    _log(SEP + "\n")

    if not os.path.isfile(ARCHIVO_INGRESO):
        _log("[ERROR] No se encontro INGRESO_MASIVO:\n  {}".format(ARCHIVO_INGRESO))
        return
    if not os.path.isdir(CARPETA_TIENDAS):
        _log("[ERROR] No se encontro carpeta tiendas:\n  {}".format(CARPETA_TIENDAS))
        return

    _log("Cargando INGRESO_MASIVO...")
    # Advertir si el INGRESO_MASIVO esta abierto en Excel
    if _archivo_bloqueado(ARCHIVO_INGRESO):
        _log("[AVISO] INGRESO_MASIVO esta abierto en Excel.")
        _log("        Los resultados no podran guardarse al terminar.")
        _log("        Cierra el archivo antes de ejecutar para evitar errores.")
    try:
        wb_origen = load_workbook(ARCHIVO_INGRESO, data_only=False)
    except Exception as e:
        _log("[ERROR] {}".format(e))
        return

    if HOJA_ORIGEN not in wb_origen.sheetnames:
        _log("[ERROR] No se encontro hoja '{}'".format(HOJA_ORIGEN))
        return

    ws_origen = wb_origen[HOJA_ORIGEN]

    mapa_tiendas = {}
    if HOJA_TIENDAS in wb_origen.sheetnames:
        mapa_tiendas = logica.cargar_mapa_tiendas(
            wb_origen[HOJA_TIENDAS], FILA_TIENDAS_INI)
        _log("[OK] Hoja TIENDAS: {} variantes".format(len(mapa_tiendas)))
    else:
        _log("[AVISO] Sin hoja TIENDAS")

    # --- Leer OMISIONES y BLACKLIST del mismo Excel (compartido en OneDrive) ---
    _crear_hoja_omisiones(wb_origen)
    _crear_hoja_blacklist(wb_origen)
    omisiones = cargar_omisiones(wb_origen)
    blacklist  = cargar_blacklist(wb_origen)
    _log("[OK] Omisiones: {} entradas".format(len(omisiones)))
    _log("[OK] Blacklist: {} tiendas bloqueadas".format(len(blacklist)))

    _log("\nIndexando carpeta de tiendas...")
    indice, indice_sin_tipo = indexar_carpeta(CARPETA_TIENDAS)
    _log("[OK] {} archivos encontrados ({} claves sin sufijo)".format(
        len(indice), len(indice_sin_tipo)))

    # Advertir bloqueados
    bloqueados = _verificar_bloqueados(indice)
    if bloqueados:
        _log("\n[AVISO] Archivos abiertos en Excel (pueden fallar al guardar):")
        for b in bloqueados:
            _log("  - {}".format(b))
        _log("")

    import indexar as _idx
    indice_cole = _idx.cargar_cache()
    if indice_cole:
        _log("[OK] Cache col E: {} nombres".format(len(indice_cole)))
    else:
        _log("[AVISO] Sin cache col E.")
    _log("")

    ultima_fila = ws_origen.max_row

    # Emitir total para la UI
    print("TOTAL_FILAS:{}".format(ultima_fila - FILA_INICIO + 1), flush=True)

    # ------------------------------------------------------------------
    # PASO 1: Pre-leer TODAS las filas en memoria (una sola pasada)
    # ------------------------------------------------------------------
    t0 = time.time()
    COL_MAX_LEER = max(COL_COMENTARIO, COL_DATOS_INI + 6) + 1
    filas_data = {}
    for row in ws_origen.iter_rows(min_row=FILA_INICIO, max_row=ultima_fila,
                                    min_col=1, max_col=COL_MAX_LEER,
                                    values_only=True):
        fila_num = FILA_INICIO + len(filas_data)
        filas_data[fila_num] = row
    _log("Leidas {} filas en {:.2f}s".format(len(filas_data), time.time() - t0))

    # ------------------------------------------------------------------
    # PASO 2: Clasificar filas por tienda en memoria (sin abrir libros)
    # ------------------------------------------------------------------
    t0 = time.time()
    grupos, faltas_clasificacion, bloqueadas_clasificacion = _clasificar_filas(
        filas_data, indice, indice_sin_tipo, mapa_tiendas, indice_cole,
        ws_origen, ultima_fila, omisiones, blacklist)

    n_tiendas = len(grupos)
    n_filas   = sum(len(v) for v in grupos.values())
    _log("Clasificadas: {} tiendas, {} filas validas, {} faltas, {} bloqueadas en {:.2f}s".format(
        n_tiendas, n_filas, len(faltas_clasificacion),
        len(bloqueadas_clasificacion), time.time() - t0))
    _log("")

    # ------------------------------------------------------------------
    # PASO 3: Procesar tienda por tienda (ya agrupadas)
    # ------------------------------------------------------------------
    cache   = CacheLibros()
    cache.iniciar_sesion_backup()
    c_listo = c_dup = 0
    filas_falta = [(fr, fv) for fr, fv, _ in faltas_clasificacion]
    c_bloq      = len(bloqueadas_clasificacion)

    # Logging de faltas de clasificacion
    for fila_real, fila_vals, motivo in faltas_clasificacion:
        if motivo != "sin_tienda":
            _log("  FALTA  fila {}: '{}'".format(fila_real, motivo))

    tiendas_procesadas = 0
    for clave_archivo, filas_grupo in grupos.items():

        # Control de cache — guardar lote si se llena
        if cache.count() >= MAX_LIBROS and not cache.tiene(clave_archivo):
            _log("\n  [LOTE] Guardando {} libros...".format(MAX_LIBROS))
            cache.cerrar_todos()

        listo, dup, faltas_grupo = _procesar_grupo(
            clave_archivo, filas_grupo, cache, indice, indice_cole, ws_origen)

        c_listo += listo
        c_dup   += dup
        filas_falta.extend([(fr, fv) for fr, fv, _ in faltas_grupo])
        tiendas_procesadas += 1

        if tiendas_procesadas % 10 == 0:
            _log("  [{}/{}] tiendas procesadas...".format(
                tiendas_procesadas, n_tiendas))

    # ------------------------------------------------------------------
    # PASO 4: Guardar todos los libros
    # ------------------------------------------------------------------
    _log("\nGuardando archivos de tiendas...")
    resumen_tiendas = cache.get_resumen()
    # Capturar lista de archivos modificados ANTES de cerrar (cerrar_todos limpia la lista)
    _archivos_modificados_bak = list(cache._archivos_originales)
    _sesion_ts_bak            = cache._sesion_ts
    _carpeta_bak_bak          = cache._carpeta_bak
    cache.cerrar_todos()

    c_falta = len(filas_falta)

    # ------------------------------------------------------------------
    # PASO 5: Actualizar hoja FALTA (solo si ya existe en el archivo)
    # ------------------------------------------------------------------
    if filas_falta:
        HOJA_FALTA = "FALTA"
        if HOJA_FALTA in wb_origen.sheetnames:
            _log("\nActualizando hoja FALTA...")
            ws_falta = wb_origen[HOJA_FALTA]
            # MEJORA: no borrar — preservar historial anterior,
            # agregar nuevas faltas debajo de las existentes
            # Encontrar la primera fila vacía desde fila 4
            _fila_falta_inicio = 4
            for _rf in range(4, ws_falta.max_row + 2):
                if all(ws_falta.cell(row=_rf, column=c).value is None
                       for c in range(1, 10)):
                    _fila_falta_inicio = _rf
                    break
        else:
            ws_falta = None
            _log("\n[AVISO] Hoja FALTA no existe en el archivo — se omite escritura.")

        if ws_falta is not None:
            # Indice id→fila para copiar estilos
            indice_id_fila = {}
            for fr in range(FILA_INICIO, ws_origen.max_row + 1):
                id_val = ws_origen.cell(row=fr, column=6).value
                res    = str(ws_origen.cell(row=fr, column=COL_RESULTADO).value or "").strip().upper()
                if id_val and res == "FALTA":
                    clave_id = str(id_val)
                    if clave_id not in indice_id_fila:
                        indice_id_fila[clave_id] = fr

            ri = _fila_falta_inicio
            for fila_real, fila_vals in filas_falta:
                id_buscar        = fila_vals[5] if len(fila_vals) > 5 else None
                fila_origen_real = indice_id_fila.get(str(id_buscar)) if id_buscar else None
                for ci, valor in enumerate(fila_vals, start=1):
                    celda_dest = ws_falta.cell(row=ri, column=ci)
                    celda_dest.value = valor
                    if fila_origen_real:
                        try:
                            celda_orig = ws_origen.cell(row=fila_origen_real, column=ci)
                            if celda_orig.has_style:
                                celda_dest.font      = copy.copy(celda_orig.font)
                                celda_dest.fill      = copy.copy(celda_orig.fill)
                                celda_dest.border    = copy.copy(celda_orig.border)
                                celda_dest.alignment = copy.copy(celda_orig.alignment)
                                celda_dest.number_format = celda_orig.number_format
                        except Exception:
                            pass
                ri += 1
            _log("[OK] {} filas en hoja FALTA".format(len(filas_falta)))

    # ------------------------------------------------------------------
    # PASO 6: Guardar INGRESO_MASIVO
    # ------------------------------------------------------------------
    _log("\nGuardando INGRESO_MASIVO...")

    # Verificar si el archivo esta bloqueado (abierto en Excel)
    if _archivo_bloqueado(ARCHIVO_INGRESO):
        _log("[AVISO] INGRESO_MASIVO esta abierto en Excel.")
        _log("        Cerralo para que se guarden los resultados.")

    import io as _io
    buf = _io.BytesIO()
    wb_origen.save(buf)
    wb_origen.close()
    contenido = buf.getvalue()

    guardado_ingreso = False
    for intento in range(1, MAX_REINTENTOS + 1):
        try:
            if contenido:
                with open(ARCHIVO_INGRESO, 'wb') as f:
                    f.write(contenido)
                _forzar_sync_onedrive(ARCHIVO_INGRESO)
                _log("[OK] INGRESO_MASIVO guardado")
                guardado_ingreso = True
            break
        except PermissionError:
            if intento < MAX_REINTENTOS:
                _log("  [Reintento {}/{}] INGRESO_MASIVO bloqueado, esperando {}s..."
                     "  (cierra el archivo en Excel)".format(
                         intento, MAX_REINTENTOS, ESPERA_REINT))
                time.sleep(ESPERA_REINT)
            else:
                _log("[ERROR] No se pudo guardar INGRESO_MASIVO despues de {} intentos.".format(
                    MAX_REINTENTOS))
                _log("        Los resultados LISTO/FALTA/DUP NO fueron escritos.")
        except Exception as e:
            _log("[ERROR] INGRESO_MASIVO: {}".format(e))
            break

    if not guardado_ingreso:
        _log("[AVISO] Cierra el INGRESO_MASIVO en Excel y vuelve a ejecutar.")

    # Guardar registro de deshacer
    if _archivos_modificados_bak:
        try:
            _deshacer.guardar_registro(
                timestamp=_sesion_ts_bak,
                carpeta_sesion=_carpeta_bak_bak,
                archivos_modificados=_archivos_modificados_bak,
                c_listo=c_listo, c_dup=c_dup, c_falta=c_falta
            )
            _log("[OK] Registro de deshacer guardado ({} archivo(s))".format(
                len(_archivos_modificados_bak)))
        except Exception as _e:
            _log("[WARN] No se pudo guardar el registro de deshacer: {}".format(_e))
    else:
        _log("[INFO] No se modificó ningún archivo — nada que deshacer.")

    t_total = time.time() - t_inicio

    # Resumen final
    _log("\n" + SEP)
    _log("  PROCESO COMPLETADO")
    _log(SEP)
    _log("  LISTO : {}".format(c_listo))
    _log("  FALTA : {}".format(c_falta))
    _log("  DUP   : {}".format(c_dup))
    _log("  BLOQ  : {}".format(c_bloq))
    _log("  Tiempo: {:.1f}s".format(t_total))

    if resumen_tiendas:
        _log("\n  RESUMEN POR TIENDA:")
        _log("  {:<40} {:>6} {:>5}".format("TIENDA", "LISTO", "DUP"))
        _log("  " + "-" * 54)
        for clave, dat in sorted(resumen_tiendas.items(),
                                  key=lambda x: x[1].get("listo", 0), reverse=True):
            if dat.get("listo", 0) > 0 or dat.get("dup", 0) > 0:
                _log("  {:<40} {:>6} {:>5}".format(
                    dat["nombre"][:40], dat.get("listo", 0), dat.get("dup", 0)))

    _log(SEP + "\n")
    if _LOG_PATH:
        _log("Log guardado en: {}".format(_LOG_PATH))


if __name__ == "__main__":
    main()


# =============================================================================
# FUNCIÓN PÚBLICA PARA EL BOTÓN DESHACER DE LA UI
# =============================================================================
def deshacer_ultimo_proceso(callback_log=None) -> tuple:
    """
    Restaura todos los archivos del último proceso desde sus backups.
    Llamada desde el botón Deshacer de la interfaz gráfica.
    
    Returns:
        (exito: bool, mensaje: str)
    """
    puede, info = _deshacer.puede_deshacer()
    if not puede:
        return False, info
    return _deshacer.deshacer(callback_log=callback_log)


def info_ultimo_proceso() -> str:
    """
    Devuelve información del último proceso para mostrar en el tooltip
    del botón Deshacer.
    """
    puede, info = _deshacer.puede_deshacer()
    return info
